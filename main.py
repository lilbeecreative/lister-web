"""
Lister AI — FastAPI Web Dashboard
Replaces Streamlit for real-time performance.
"""
import os
import csv
import io
from datetime import datetime
from dotenv import load_dotenv
from fastapi import FastAPI, Request, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from supabase import create_client
from pydantic import BaseModel
from typing import Optional

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
print(f"Connecting to Supabase: {SUPABASE_URL}")
supabase     = create_client(SUPABASE_URL, SUPABASE_KEY)

app = FastAPI(title="Lister AI")
import os as _os
if _os.path.isdir("static"):
    app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

EBAY_DESCRIPTION = "Shipped primarily with UPS and sometimes USPS. If you have special packing or shipping needs, please send a message. This item is sold in as-is condition. The seller assumes no liability for the use, operation, or installation of this product. Due to the technical nature of this equipment, the buyer is responsible for having the item professionally inspected and installed by a certified technician prior to use."

def photo_url(photo_id: str, thumb: bool = False) -> str:
    if not photo_id or photo_id in ("", "nan", "0"):
        return ""
    if thumb:
        return f"{SUPABASE_URL}/storage/v1/render/image/public/part-photos/{photo_id}?width=500&height=500&resize=cover&quality=80"
    return f"{SUPABASE_URL}/storage/v1/object/public/part-photos/{photo_id}"

# ── PAGES ─────────────────────────────────────────────────────── #

@app.get("/auction/research", response_class=HTMLResponse)
async def auction_research_page(request: Request):
    import os
    with open(os.path.join(os.path.dirname(__file__), "templates", "auction_research.html")) as f:
        html = f.read()
    return HTMLResponse(content=html, headers={
        "Content-Security-Policy": "default-src * blob: data:; script-src * blob: data: 'unsafe-inline' 'unsafe-eval'; style-src * 'unsafe-inline'; img-src * blob: data:;"
    })

@app.get("/auction", response_class=HTMLResponse)
async def auction_page(request: Request):
    business_id = require_auth(request)
    if not business_id:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    import os
    with open(os.path.join(os.path.dirname(__file__), "templates", "auction.html")) as f:
        html = f.read()
    return HTMLResponse(content=html, headers={
        "Content-Security-Policy": "default-src * blob: data:; script-src * blob: data: 'unsafe-inline' 'unsafe-eval'; style-src * 'unsafe-inline'; img-src * blob: data:;"
    })

@app.get("/v2", response_class=HTMLResponse)
async def dashboard_v2(request: Request):
    from fastapi.responses import HTMLResponse
    import os
    with open(os.path.join(os.path.dirname(__file__), "templates", "v2.html")) as f:
        html = f.read()
    return HTMLResponse(content=html, headers={
        "Content-Security-Policy": "default-src * blob: data:; script-src * blob: data: 'unsafe-inline' 'unsafe-eval'; style-src * 'unsafe-inline'; img-src * blob: data:;"
    })
@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    with open(os.path.join(os.path.dirname(__file__), "templates", "admin.html")) as f:
        return HTMLResponse(f.read())

@app.patch("/api/admin/businesses/{business_id}")
async def admin_update_business(business_id: str, request: Request):
    try:
        body = await request.json()
        allowed = {k: v for k, v in body.items() if k in ("scan_limit", "scan_count", "is_admin")}
        supabase.table("businesses").update(allowed).eq("id", business_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.delete("/api/admin/businesses/{business_id}")
async def admin_delete_business(business_id: str):
    try:
        supabase.table("sessions").delete().eq("business_id", business_id).execute()
        supabase.table("listings").delete().eq("business_id", business_id).execute()
        supabase.table("listing_groups").delete().eq("business_id", business_id).execute()
        supabase.table("businesses").delete().eq("id", business_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/admin/businesses")
async def admin_businesses():
    try:
        biz = supabase.table("businesses").select("id,name,email,created_at,scan_count,scan_limit,is_admin").order("created_at", desc=True).execute()
        businesses = biz.data or []
        for b in businesses:
            lid = supabase.table("listings").select("id", count="exact").eq("business_id", b["id"]).execute()
            b["listing_count"] = lid.count or 0
            # Use last listing created_at as last_active
            last = supabase.table("listings").select("created_at").eq("business_id", b["id"]).order("created_at", desc=True).limit(1).execute()
            b["last_active"] = last.data[0]["created_at"] if last.data else None
            # scan_count already comes from businesses table
        return {"businesses": businesses}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/landing", response_class=HTMLResponse)
async def landing_page_route(request: Request):
    with open(os.path.join(os.path.dirname(__file__), "templates", "landing.html")) as f:
        return HTMLResponse(f.read())

@app.get("/checkout", response_class=HTMLResponse)
async def checkout_page(request: Request):
    business_id = require_auth(request)
    if not business_id:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    with open(os.path.join(os.path.dirname(__file__), "templates", "checkout.html")) as f:
        return HTMLResponse(f.read())

@app.get("/welcome", response_class=HTMLResponse)
async def welcome_page(request: Request):
    business_id = require_auth(request)
    if not business_id:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    # Skip onboarding if already done
    try:
        biz = supabase.table("businesses").select("onboarded").eq("id", business_id).limit(1).execute()
        if biz.data and biz.data[0].get("onboarded"):
            from fastapi.responses import RedirectResponse
            return RedirectResponse("/", status_code=302)
    except Exception:
        pass
    with open(os.path.join(os.path.dirname(__file__), "templates", "welcome.html")) as f:
        return HTMLResponse(f.read())


@app.post("/api/welcome/complete")
async def mark_onboarded(request: Request):
    """Called by the onboarding flow to mark user as onboarded so they don't loop back."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        supabase.table("businesses").update({"onboarded": True}).eq("id", business_id).execute()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.post("/api/checkout")
async def process_checkout(request: Request):
    import resend
    body = await request.json()
    plan = body.get("plan", "Free Trial")
    price = body.get("price", "0")
    scans = body.get("scans", "25 total scans")
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        # Update scan limit based on plan
        plan_limits = {"Free Trial": 25, "Starter": 100, "Growth": 500, "Pro": 1000}
        scan_limit = plan_limits.get(plan, 25)
        biz = supabase.table("businesses").select("name,email").eq("id", business_id).execute()
        if not biz.data:
            raise HTTPException(404, "Business not found")
        biz_data = biz.data[0]
        supabase.table("businesses").update({"scan_limit": scan_limit, "onboarded": True}).eq("id", business_id).execute()
        # Send welcome email
        resend.api_key = os.getenv("RESEND_API_KEY", "re_7V1Ykni2_FjGBPZwsgWr8ytntk3sCoG5b")
        resend.Emails.send({
            "from": "Lister AI <hello@reselljunkie.com>",
            "to": biz_data["email"],
            "subject": "Welcome to Lister AI — You're all set!",
            "html": f"""
<div style="font-family:-apple-system,sans-serif;max-width:560px;margin:0 auto;background:#060608;color:#f0f2f5;padding:40px 32px;border-radius:16px;">
  <div style="font-size:32px;font-weight:900;letter-spacing:0.05em;margin-bottom:32px;">Lister<span style="color:#e8ff47;">AI</span></div>
  <h1 style="font-size:28px;font-weight:800;margin-bottom:12px;">Welcome, {biz_data['name']}! 🎉</h1>
  <p style="color:#8892a4;font-size:15px;line-height:1.7;margin-bottom:24px;">Your account is active. You're on the <strong style="color:#e8ff47;">{plan} plan</strong> with <strong style="color:#e8ff47;">{scans}</strong>.</p>
  <a href="https://lister-web-dev-production.up.railway.app/" style="display:inline-block;background:#e8ff47;color:#000;padding:14px 32px;border-radius:8px;text-decoration:none;font-weight:700;font-size:15px;margin-bottom:32px;">Go to Dashboard →</a>
  <div style="background:#111418;border:1px solid #1a1f2e;border-radius:12px;padding:24px;margin-bottom:24px;">
    <h2 style="font-size:16px;font-weight:700;margin-bottom:16px;">Quick Start in 5 Steps</h2>
    <div style="margin-bottom:12px;"><strong style="color:#e8ff47;">1.</strong> <span style="color:#8892a4;">Open your dashboard and click the Upload tab</span></div>
    <div style="margin-bottom:12px;"><strong style="color:#e8ff47;">2.</strong> <span style="color:#8892a4;">Select New or Used condition and start a batch</span></div>
    <div style="margin-bottom:12px;"><strong style="color:#e8ff47;">3.</strong> <span style="color:#8892a4;">Take or upload photos — the AI identifies and prices automatically</span></div>
    <div style="margin-bottom:12px;"><strong style="color:#e8ff47;">4.</strong> <span style="color:#8892a4;">Review and edit listings in the dashboard</span></div>
    <div><strong style="color:#e8ff47;">5.</strong> <span style="color:#8892a4;">Export to eBay using the CSV button in the sidebar</span></div>
  </div>
  <p style="color:#5a6478;font-size:12px;">Need help? Reply to this email and we'll get back to you.</p>
</div>
"""
        })
        return {"ok": True}
    except Exception as e:
        print(f"Checkout error: {e}")
        return {"ok": True}  # Still redirect even if email fails

# ── EBAY OAUTH & API ────────────────────────────────────────────
import base64, urllib.parse, requests as _req

EBAY_APP_ID = os.getenv("EBAY_APP_ID", "")
EBAY_CERT_ID = os.getenv("EBAY_CERT_ID", "")
EBAY_RUNAME = os.getenv("EBAY_RUNAME", "")
EBAY_ENV = os.getenv("EBAY_ENV", "sandbox")
EBAY_API_BASE = "https://api.sandbox.ebay.com" if EBAY_ENV == "sandbox" else "https://api.ebay.com"
EBAY_AUTH_BASE = "https://auth.sandbox.ebay.com" if EBAY_ENV == "sandbox" else "https://auth.ebay.com"
EBAY_SCOPES = "https://api.ebay.com/oauth/api_scope/sell.inventory https://api.ebay.com/oauth/api_scope/sell.account"

@app.get("/ebay/connect")
async def ebay_connect(request: Request):
    business_id = require_auth(request)
    if not business_id:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    auth_url = (f"{EBAY_AUTH_BASE}/oauth2/authorize?client_id={EBAY_APP_ID}&response_type=code&redirect_uri={EBAY_RUNAME}&scope={urllib.parse.quote(EBAY_SCOPES)}&state={business_id}")
    from fastapi.responses import RedirectResponse
    return RedirectResponse(auth_url, status_code=302)

@app.get("/ebay/callback")
async def ebay_callback(request: Request, code: str = "", state: str = ""):
    from fastapi.responses import HTMLResponse, RedirectResponse
    if not code or not state:
        return HTMLResponse("<h1>eBay Connection Failed</h1><p>Missing code or state. <a href='/'>Go back</a></p>")
    creds = base64.b64encode(f"{EBAY_APP_ID}:{EBAY_CERT_ID}".encode()).decode()
    try:
        r = _req.post(f"{EBAY_API_BASE}/identity/v1/oauth2/token", headers={"Authorization": f"Basic {creds}", "Content-Type": "application/x-www-form-urlencoded"}, data={"grant_type": "authorization_code", "code": code, "redirect_uri": EBAY_RUNAME})
        data = r.json()
        if "access_token" not in data:
            return HTMLResponse(f"<h1>eBay Connection Failed</h1><pre>{data}</pre><a href='/'>Go back</a>")
        from datetime import datetime, timedelta
        expires_at = (datetime.utcnow() + timedelta(seconds=data.get("expires_in", 7200))).isoformat()
        existing = supabase.table("ebay_tokens").select("id").eq("business_id", state).execute()
        token_row = {"business_id": state, "access_token": data["access_token"], "refresh_token": data.get("refresh_token", ""), "expires_at": expires_at, "environment": EBAY_ENV, "updated_at": datetime.utcnow().isoformat()}
        if existing.data:
            supabase.table("ebay_tokens").update(token_row).eq("business_id", state).execute()
        else:
            supabase.table("ebay_tokens").insert(token_row).execute()
        return RedirectResponse("/?ebay_connected=1", status_code=302)
    except Exception as e:
        return HTMLResponse(f"<h1>eBay Error</h1><pre>{e}</pre><a href='/'>Go back</a>")

@app.get("/ebay/declined")
async def ebay_declined():
    from fastapi.responses import HTMLResponse
    return HTMLResponse("<h1>eBay Connection Cancelled</h1><a href='/'>Go back to dashboard</a>")

@app.get("/api/ebay/status")
async def ebay_status(request: Request):
    business_id = require_auth(request)
    if not business_id:
        return {"connected": False}
    try:
        res = supabase.table("ebay_tokens").select("expires_at,environment").eq("business_id", business_id).execute()
        if res.data:
            return {"connected": True, "environment": res.data[0]["environment"], "expires_at": res.data[0]["expires_at"]}
        return {"connected": False}
    except Exception:
        return {"connected": False}

@app.post("/api/ebay/disconnect")
async def ebay_disconnect(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    supabase.table("ebay_tokens").delete().eq("business_id", business_id).execute()
    return {"ok": True}

# ── SHOPIFY OAUTH ────────────────────────────────────────────
SHOPIFY_API_KEY = os.getenv("SHOPIFY_API_KEY", "")
SHOPIFY_API_SECRET = os.getenv("SHOPIFY_API_SECRET", "")
SHOPIFY_SCOPES = os.getenv("SHOPIFY_SCOPES", "write_products,read_products")

@app.get("/shopify/connect")
async def shopify_connect(request: Request, shop: str = ""):
    business_id = require_auth(request)
    if not business_id:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    if not shop:
        from fastapi.responses import HTMLResponse
        return HTMLResponse("""
<html><body style="background:#060608;color:#f0f2f5;font-family:sans-serif;padding:60px 20px;text-align:center;">
<h2>Connect Your Shopify Store</h2>
<form method="GET" action="/shopify/connect" style="max-width:400px;margin:30px auto;">
  <input name="shop" placeholder="your-store.myshopify.com" required style="width:100%;padding:12px;background:#0f1117;border:1px solid #1e2330;border-radius:8px;color:#fff;font-size:14px;margin-bottom:12px;">
  <button type="submit" style="width:100%;background:#22c55e;color:#000;padding:12px;border:none;border-radius:8px;font-weight:700;cursor:pointer;">Connect</button>
</form>
<a href="/" style="color:#5a6478;font-size:13px;">← Back</a>
</body></html>""")
    if not shop.endswith(".myshopify.com"):
        shop = shop.replace("https://","").replace("http://","").rstrip("/")
        if not shop.endswith(".myshopify.com"):
            shop = shop + ".myshopify.com"
    auth_url = f"https://{shop}/admin/oauth/authorize?client_id={SHOPIFY_API_KEY}&scope={SHOPIFY_SCOPES}&redirect_uri=https://lister-web-dev-production.up.railway.app/shopify/callback&state={business_id}__{shop}"
    from fastapi.responses import RedirectResponse
    return RedirectResponse(auth_url, status_code=302)

@app.get("/shopify/callback")
async def shopify_callback(request: Request, code: str = "", state: str = "", shop: str = ""):
    from fastapi.responses import HTMLResponse, RedirectResponse
    if not code or not state:
        return HTMLResponse("<h1>Shopify Connection Failed</h1><a href='/'>Go back</a>")
    parts = state.split("__")
    if len(parts) != 2:
        return HTMLResponse("<h1>Invalid state</h1><a href='/'>Go back</a>")
    business_id, shop_domain = parts
    try:
        r = _req.post(f"https://{shop_domain}/admin/oauth/access_token", json={
            "client_id": SHOPIFY_API_KEY,
            "client_secret": SHOPIFY_API_SECRET,
            "code": code
        })
        data = r.json()
        if "access_token" not in data:
            return HTMLResponse(f"<h1>Failed</h1><pre>{data}</pre><a href='/'>Go back</a>")
        from datetime import datetime
        existing = supabase.table("shopify_tokens").select("id").eq("business_id", business_id).execute()
        token_row = {"business_id": business_id, "shop_domain": shop_domain, "access_token": data["access_token"], "scopes": data.get("scope", ""), "updated_at": datetime.utcnow().isoformat()}
        if existing.data:
            supabase.table("shopify_tokens").update(token_row).eq("business_id", business_id).execute()
        else:
            supabase.table("shopify_tokens").insert(token_row).execute()
        return RedirectResponse("/?shopify_connected=1", status_code=302)
    except Exception as e:
        return HTMLResponse(f"<h1>Error</h1><pre>{e}</pre><a href='/'>Go back</a>")

@app.get("/api/shopify/status")
async def shopify_status(request: Request):
    business_id = require_auth(request)
    if not business_id:
        return {"connected": False}
    try:
        res = supabase.table("shopify_tokens").select("shop_domain").eq("business_id", business_id).execute()
        if res.data:
            return {"connected": True, "shop_domain": res.data[0]["shop_domain"]}
        return {"connected": False}
    except Exception:
        return {"connected": False}

@app.post("/api/shopify/disconnect")
async def shopify_disconnect(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    supabase.table("shopify_tokens").delete().eq("business_id", business_id).execute()
    return {"ok": True}

@app.post("/api/listings/{listing_id}/report-bad-scan")
async def report_bad_scan(listing_id: int, request: Request):
    from datetime import datetime
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        reason = body.get("reason", "")
        # Get listing
        l = supabase.table("listings").select("*").eq("id", listing_id).eq("business_id", business_id).execute()
        if not l.data:
            raise HTTPException(404, "Listing not found")
        listing = l.data[0]
        # Get business info
        biz = supabase.table("businesses").select("name,email").eq("id", business_id).execute()
        biz_data = biz.data[0] if biz.data else {}
        # Get all photos and create signed URLs (work even if bucket is private)
        photo_ids = []
        if listing.get("photo_id"):
            gp = supabase.table("group_photos").select("group_id").eq("photo_id", listing["photo_id"]).execute()
            if gp.data:
                gid = gp.data[0]["group_id"]
                all_gp = supabase.table("group_photos").select("photo_id").eq("group_id", gid).execute()
                for row in (all_gp.data or []):
                    photo_ids.append(row['photo_id'])
            if not photo_ids:
                photo_ids.append(listing['photo_id'])
        # Generate signed URLs (1 day expiry)
        photo_urls = []
        for pid in photo_ids:
            try:
                signed = supabase.storage.from_("part-photos").create_signed_url(pid, 86400)
                if signed and "signedURL" in signed:
                    photo_urls.append(signed["signedURL"])
                elif signed and "signedUrl" in signed:
                    photo_urls.append(signed["signedUrl"])
                else:
                    photo_urls.append(f"{SUPABASE_URL}/storage/v1/object/public/part-photos/{pid}")
            except Exception:
                photo_urls.append(f"{SUPABASE_URL}/storage/v1/object/public/part-photos/{pid}")
        # Save to bad_scan_reports table for admin dashboard
        try:
            supabase.table("bad_scan_reports").insert({
                "reporter_email": biz_data.get("email", ""),
                "business_id": business_id,
                "group_id": None,
                "photo_ids": photo_ids if photo_ids else [],
                "item_card": listing,
                "user_note": reason,
            }).execute()
        except Exception as _e:
            print(f"bad_scan_reports insert failed: {_e}")
        # Send email via Resend
        import resend
        resend.api_key = os.getenv("RESEND_API_KEY", "")
        print(f"📸 Bad scan photos: {photo_urls}")
        # Build public URLs for email reliability (signed URLs can expire/break)
        public_urls = [f"{SUPABASE_URL}/storage/v1/object/public/part-photos/{pid}" for pid in photo_ids]
        photo_html = "".join([
            f'<div style="margin:8px 0;"><img src="{u}" style="max-width:300px;border-radius:8px;border:1px solid #ddd;display:block;margin-bottom:6px;"/><a href="{u}" style="color:#3b82f6;font-size:11px;word-break:break-all;">{u}</a></div>'
            for u in public_urls[:5]
        ])
        email_html = f"""
<div style="font-family:-apple-system,sans-serif;max-width:640px;margin:0 auto;padding:24px;background:#0a0c10;color:#f0f2f5;">
  <h1 style="color:#ef4444;margin-bottom:8px;font-size:22px;">⚠️ Bad Scan Reported</h1>
  <p style="color:#8892a4;font-size:13px;margin-bottom:24px;">A user flagged this listing as a bad scan and needs review.</p>
  
  <div style="background:#111418;border:1px solid #1a1f2e;border-radius:10px;padding:20px;margin-bottom:16px;">
    <h2 style="font-size:14px;color:#e8ff47;margin-bottom:12px;text-transform:uppercase;letter-spacing:0.08em;">User Info</h2>
    <p style="margin:4px 0;"><strong>Business:</strong> {biz_data.get('name','Unknown')}</p>
    <p style="margin:4px 0;"><strong>Email:</strong> {biz_data.get('email','Unknown')}</p>
    <p style="margin:4px 0;"><strong>Business ID:</strong> {business_id}</p>
    <p style="margin:4px 0;"><strong>Listing ID:</strong> {listing_id}</p>
    <p style="margin:4px 0;"><strong>Photo ID(s):</strong> {", ".join(photo_ids) if photo_ids else "(none)"}</p>
    <p style="margin:4px 0;"><strong>Listing Created:</strong> {listing.get('created_at','Unknown')}</p>
    <p style="margin:4px 0;"><strong>Reported At:</strong> {datetime.utcnow().isoformat()}Z</p>
  </div>
  
  <div style="background:#111418;border:1px solid #1a1f2e;border-radius:10px;padding:20px;margin-bottom:16px;">
    <h2 style="font-size:14px;color:#e8ff47;margin-bottom:12px;text-transform:uppercase;letter-spacing:0.08em;">User Reason</h2>
    <p style="font-style:italic;color:#8892a4;">{reason or '(no reason given)'}</p>
  </div>
  
  <div style="background:#111418;border:1px solid #1a1f2e;border-radius:10px;padding:20px;margin-bottom:16px;">
    <h2 style="font-size:14px;color:#e8ff47;margin-bottom:12px;text-transform:uppercase;letter-spacing:0.08em;">AI Generated Data</h2>
    <p style="margin:4px 0;"><strong>Title:</strong> {listing.get('title','')}</p>
    <p style="margin:4px 0;"><strong>Description:</strong> {(listing.get('description') or '')[:200]}</p>
    <p style="margin:4px 0;"><strong>Price (Used):</strong> ${listing.get('price_used',0)}</p>
    <p style="margin:4px 0;"><strong>Price (New):</strong> ${listing.get('price_new',0)}</p>
    <p style="margin:4px 0;"><strong>Condition:</strong> {listing.get('condition','')}</p>
    <p style="margin:4px 0;"><strong>Quantity:</strong> {listing.get('quantity',1)}</p>
    <p style="margin:4px 0;"><strong>eBay Category:</strong> {listing.get('ebay_category','')} (ID: {listing.get('ebay_category_id','')})</p>
  </div>
  
  <div style="background:#111418;border:1px solid #1a1f2e;border-radius:10px;padding:20px;">
    <h2 style="font-size:14px;color:#e8ff47;margin-bottom:12px;text-transform:uppercase;letter-spacing:0.08em;">Photos ({len(photo_urls)})</h2>
    <div>{photo_html}</div>
  </div>
</div>
"""
        try:
            resend.Emails.send({
                "from": "Lister AI Reports <reports@reselljunkie.com>",
                "to": "sebastian@lilbeecreative.com",
                "subject": f"Bad Scan Report — {biz_data.get('name','Unknown')} — Listing #{listing_id}",
                "html": email_html
            })
        except Exception as email_err:
            # Try fallback sender
            resend.Emails.send({
                "from": "Lister AI <onboarding@resend.dev>",
                "to": "sebastian@lilbeecreative.com",
                "subject": f"Bad Scan Report — Listing #{listing_id}",
                "html": email_html
            })
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"❌ Bad scan report error: {e}")
        raise HTTPException(500, str(e))

@app.post("/api/shopify/push/{listing_id}")
async def push_to_shopify(listing_id: int, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        # Get token
        tok = supabase.table("shopify_tokens").select("*").eq("business_id", business_id).execute()
        if not tok.data:
            raise HTTPException(400, "Shopify not connected")
        shop = tok.data[0]["shop_domain"]
        token = tok.data[0]["access_token"]
        # Get listing
        l = supabase.table("listings").select("*").eq("id", listing_id).eq("business_id", business_id).execute()
        if not l.data:
            raise HTTPException(404, "Listing not found")
        listing = l.data[0]
        # Get all photos via group
        photo_urls = []
        if listing.get("photo_id"):
            gp = supabase.table("group_photos").select("group_id").eq("photo_id", listing["photo_id"]).execute()
            if gp.data:
                gid = gp.data[0]["group_id"]
                all_gp = supabase.table("group_photos").select("photo_id").eq("group_id", gid).execute()
                for row in (all_gp.data or []):
                    photo_urls.append(f"{SUPABASE_URL}/storage/v1/object/public/part-photos/{row['photo_id']}")
            if not photo_urls:
                photo_urls.append(f"{SUPABASE_URL}/storage/v1/object/public/part-photos/{listing['photo_id']}")
        # Create product on Shopify
        product_data = {
            "product": {
                "title": listing.get("title", "Untitled"),
                "body_html": listing.get("description", ""),
                "vendor": "Lister AI",
                "product_type": listing.get("ebay_category", ""),
                "status": "draft",
                "variants": [{
                    "price": str(listing.get("price", 0)),
                    "inventory_quantity": listing.get("quantity", 1),
                    "inventory_management": "shopify",
                }],
                "images": [{"src": url} for url in photo_urls[:10]],
            }
        }
        r = _req.post(
            f"https://{shop}/admin/api/2024-01/products.json",
            headers={"X-Shopify-Access-Token": token, "Content-Type": "application/json"},
            json=product_data,
            timeout=30
        )
        if r.status_code not in (200, 201):
            raise HTTPException(500, f"Shopify error: {r.text}")
        result = r.json()
        product_id = result.get("product", {}).get("id")
        # Update listing with shopify_product_id
        supabase.table("listings").update({"shopify_product_id": str(product_id)}).eq("id", listing_id).execute()
        return {"ok": True, "product_id": product_id, "shop": shop}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/paywall", response_class=HTMLResponse)
async def paywall_page(request: Request):
    with open(os.path.join(os.path.dirname(__file__), "templates", "paywall.html")) as f:
        return HTMLResponse(f.read())

@app.get("/internal", response_class=HTMLResponse)
async def internal_page(request: Request):
    with open(os.path.join(os.path.dirname(__file__), "templates", "internal.html")) as f:
        return HTMLResponse(f.read())

@app.get("/team", response_class=HTMLResponse)
async def team_portal(request: Request):
    with open(os.path.join(os.path.dirname(__file__), "templates", "team_portal.html")) as f:
        return HTMLResponse(f.read())

@app.get("/portal", response_class=HTMLResponse)
async def portal(request: Request):
    import os
    with open(os.path.join(os.path.dirname(__file__), "templates", "portal.html")) as f:
        html = f.read()
    from fastapi.responses import HTMLResponse
    return HTMLResponse(content=html, headers={
        "Content-Security-Policy": "default-src * blob: data:; script-src * blob: data: 'unsafe-inline' 'unsafe-eval'; style-src * 'unsafe-inline'; img-src * blob: data:;"
    })


def require_auth(request: Request):
    """Returns business_id if authenticated, else None."""
    token = request.cookies.get("session_id")
    if not token:
        return None
    try:
        res = supabase.table("sessions").select("business_id").eq("token", token).execute()
        if res.data:
            return res.data[0]["business_id"]
    except Exception:
        pass
    return None

def get_business_info(request: Request):
    """Returns (business_id, is_admin) or (None, False)."""
    token = request.cookies.get("session_id")
    if not token:
        return None, False
    try:
        res = supabase.table("sessions").select("business_id").eq("token", token).execute()
        if not res.data:
            return None, False
        bid = res.data[0]["business_id"]
        biz = supabase.table("businesses").select("is_admin").eq("id", bid).execute()
        is_admin = bool(biz.data[0].get("is_admin") or False) if biz.data else False
        return bid, is_admin
    except Exception:
        pass
    return None, False

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    business_id, is_admin = get_business_info(request)
    if not business_id:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/login", status_code=302)
    biz = supabase.table("businesses").select("name,scan_count,scan_limit").eq("id", business_id).execute()
    biz_data = biz.data[0] if biz.data else {}
    return templates.TemplateResponse("index.html", {
        "request": request,
        "is_admin": is_admin,
        "biz_name": biz_data.get("name", ""),
        "scan_count": biz_data.get("scan_count", 0) or 0,
        "scan_limit": biz_data.get("scan_limit", 25) or 25,
    })

# ── API: LISTINGS ─────────────────────────────────────────────── #

@app.get("/api/listings")
async def get_listings(request: Request):
    business_id = require_auth(request)
    from fastapi.responses import JSONResponse
    if not business_id:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    try:
        res = supabase.table("listings")\
            .select("*")\
            .eq("business_id", business_id)\
            .neq("status", "archived")\
            .order("created_at", desc=True)\
            .execute()
        listings = res.data or []

        # Batch fetch all group photos for these listings
        primary_pids = [str(l.get("photo_id") or "") for l in listings if l.get("photo_id")]
        group_photo_map = {}  # photo_id -> [all photo_ids in same group]
        if primary_pids:
            try:
                gp_res = supabase.table("group_photos")                    .select("group_id, photo_id")                    .in_("photo_id", primary_pids[:100])                    .execute()
                # Map primary photo -> group_id
                pid_to_gid = {row["photo_id"]: row["group_id"] for row in (gp_res.data or [])}
                group_ids = list(set(pid_to_gid.values()))
                if group_ids:
                    all_gp = supabase.table("group_photos")                        .select("group_id, photo_id")                        .in_("group_id", group_ids)                        .execute()
                    # Build group_id -> [photo_ids]
                    gid_to_photos = {}
                    for row in (all_gp.data or []):
                        gid_to_photos.setdefault(row["group_id"], []).append(row["photo_id"])
                    # Map primary photo_id -> all photos in its group
                    for pid, gid in pid_to_gid.items():
                        group_photo_map[pid] = gid_to_photos.get(gid, [pid])
            except Exception as search_err:
                print(f"   Search grounding failed: {search_err}")


        for l in listings:
            pid = str(l.get("photo_id") or "")
            all_photos = group_photo_map.get(pid, [pid] if pid else [])
            l["thumb_url"]  = photo_url(pid, thumb=True)
            l["full_url"]   = photo_url(pid)
            l["all_photos"] = [{"thumb": photo_url(p, thumb=True), "full": photo_url(p)} for p in all_photos if p]
            # Coerce types
            l["price"]      = float(l.get("price") or 0)
            l["price_used"] = float(l.get("price_used") or 0)
            l["price_new"]  = float(l.get("price_new") or 0)
            l["quantity"]   = int(l.get("quantity") or 1)
            # Normalize condition — default to "used" if blank/null
            cond = str(l.get("condition") or "").strip().lower()
            l["condition"] = cond if cond in ("new", "used") else "used"
            # Normalize listing_type — items from batch upload are not auctions
            lt = str(l.get("listing_type") or "").strip().lower()
            if lt not in ("auction", "fixed"):
                l["listing_type"] = "fixed"
        return JSONResponse(listings)
    except Exception as e:
        raise HTTPException(500, str(e))

class UpdateField(BaseModel):
    field: str
    value: object




def get_ebay_token(business_id: str) -> str:
    """Get valid eBay access token, refreshing if needed."""
    import requests as _req2
    from datetime import datetime
    res = supabase.table("ebay_tokens").select("*").eq("business_id", business_id).limit(1).execute()
    if not res.data:
        raise Exception("eBay not connected")
    tok = res.data[0]
    # Check if expired
    try:
        expires_str = tok["expires_at"].replace("Z","").replace("+00:00","")
        # Strip timezone if present
        if "+" in expires_str:
            expires_str = expires_str.split("+")[0]
        expires_at = datetime.fromisoformat(expires_str)
        if (expires_at - datetime.utcnow()).total_seconds() < 300:
            # Refresh token
            creds = base64.b64encode(f"{EBAY_APP_ID}:{EBAY_CERT_ID}".encode()).decode()
            r = _req2.post(f"{EBAY_API_BASE}/identity/v1/oauth2/token",
                headers={"Authorization": f"Basic {creds}", "Content-Type": "application/x-www-form-urlencoded"},
                data={"grant_type": "refresh_token", "refresh_token": tok["refresh_token"]})
            data = r.json()
            if "access_token" in data:
                from datetime import timedelta
                new_expires = (datetime.utcnow() + timedelta(seconds=data.get("expires_in", 7200))).isoformat()
                supabase.table("ebay_tokens").update({
                    "access_token": data["access_token"],
                    "expires_at": new_expires
                }).eq("business_id", business_id).execute()
                return data["access_token"]
    except Exception as _e:
        print(f"Token refresh warning: {_e}")
    return tok["access_token"]


@app.post("/api/ebay/submit-listings")
async def submit_listings_to_ebay(request: Request):
    """Submit selected listings to eBay as draft listings using Inventory API."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        listing_ids = body.get("listing_ids", [])
        modal_items = {str(item.get("id")): item for item in body.get("items", [])}
        print(f"[eBay] Submit: {len(listing_ids)} ids, {len(modal_items)} modal items")
        print(f"[eBay] First modal item: {body.get('items', [{}])[0] if body.get('items') else 'EMPTY'}")
        if not listing_ids:
            raise HTTPException(400, "No listing IDs provided")

        token = get_ebay_token(business_id)
        env = "production" if EBAY_ENV == "production" else "sandbox"
        api_base = "https://api.ebay.com" if env == "production" else "https://api.sandbox.ebay.com"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "Content-Language": "en-US"
        }

        # Fetch listings from DB
        res = supabase.table("listings").select("*").in_("id", listing_ids).eq("business_id", business_id).execute()
        listings_data = res.data or []
        print(f"[eBay] Found {len(listings_data)} listings in DB for {len(listing_ids)} requested ids")

        results = []
        for listing in listings_data:
            try:
                lid = listing["id"]
                sku = f"LISTER-{str(lid)[:8]}"
                title = (listing.get("title") or "Item")[:80]
                price = float(listing.get("price") or 9.99)
                condition = "USED_EXCELLENT" if (listing.get("condition") or "used").lower() == "used" else "NEW"
                category_id = str(listing.get("ebay_category_id") or "99")
                print(f"[eBay] Using category_id={category_id} for {listing.get('id')}")
                description = listing.get("description") or title
                qty = int(listing.get("quantity") or 1)

                # Get photo URL
                photo_url = None
                if listing.get("photo_id"):
                    photo_res = supabase.storage.from_("part-photos").get_public_url(listing["photo_id"])
                    photo_url = photo_res

                print(f"[eBay] Processing item {lid}: title={title!r}, price={price}, sku={sku}")
                # Import requests early
                import requests as _req3

                # 1. Create/update inventory item
                # Fetch required aspects for this category from eBay
                inv_aspects = {}
                try:
                    asp_r = _req3.get(
                        f"{api_base}/commerce/taxonomy/v1/category_tree/0/get_item_aspects_for_category?category_id={category_id}",
                        headers=headers
                    )
                    print(f"[eBay] Aspect API status={asp_r.status_code}, body={asp_r.text[:300]}")
                    if asp_r.ok:
                        asp_data = asp_r.json()
                        all_aspects = asp_data.get("aspects", [])
                        print(f"[eBay] Found {len(all_aspects)} aspects from API")
                        for asp in all_aspects:
                            asp_name = asp.get("localizedAspectName")
                            constraint = asp.get("aspectConstraint", {})
                            usage = constraint.get("aspectUsage", "OPTIONAL")
                            mode = constraint.get("aspectMode", "FREE_TEXT")
                            if usage in ("REQUIRED", "RECOMMENDED"):
                                # Try to extract from title
                                title_lower = title.lower()
                                values = asp.get("aspectValues", [])
                                matched_value = None
                                for v in values[:50]:
                                    val_text = v.get("localizedValue", "")
                                    if val_text and val_text.lower() in title_lower:
                                        matched_value = val_text
                                        break
                                # If no match found, use first value or N/A
                                if matched_value:
                                    inv_aspects[asp_name] = [matched_value]
                                elif mode == "SELECTION_ONLY" and values:
                                    inv_aspects[asp_name] = [values[0].get("localizedValue", "N/A")]
                                else:
                                    inv_aspects[asp_name] = ["N/A"] if usage == "REQUIRED" else None
                                if inv_aspects.get(asp_name) is None:
                                    del inv_aspects[asp_name]
                    print(f"[eBay] Auto-aspects for {category_id}: {inv_aspects}")
                except Exception as _asp_err:
                    print(f"[eBay] Aspect fetch failed: {_asp_err}")

                # If aspects empty (e.g. permission denied), use generic defaults
                if not inv_aspects:
                    inv_aspects = {
                        "Brand": ["Unbranded"], "Type": ["Other"], "Model": ["Generic"],
                        "MPN": ["Does Not Apply"], "Color": ["Multicolor"], "Size": ["Standard"],
                        "Size Type": ["Regular"], "Material": ["Mixed Materials"],
                        "Department": ["Unisex Adult"], "Style": ["Casual"], "Theme": ["General"],
                        "Movie/TV Title": ["N/A"], "Genre": ["Other"], "Format": ["Standard"]
                    }
                    print(f"[eBay] Using generic fallback aspects")

                # Override with any user-provided aspects from modal
                user_aspects = modal.get("aspects") if 'modal' in dir() else {}
                if user_aspects:
                    for k, v in user_aspects.items():
                        if v:
                            inv_aspects[k] = [str(v)] if not isinstance(v, list) else v

                inv_payload = {
                    "availability": {"shipToLocationAvailability": {"quantity": qty}},
                    "condition": condition,
                    "product": {
                        "title": title,
                        "description": description if len(description) >= 30 else (description + " - Listed via Lister AI"),
                        "aspects": inv_aspects,
                        "country": "US"
                    }
                }
                if photo_url and photo_url.startswith("https://"):
                    inv_payload["product"]["imageUrls"] = [photo_url]
                print(f"[eBay] Payload for {sku}: {inv_payload}")

                import requests as _req3
                inv_r = _req3.put(
                    f"{api_base}/sell/inventory/v1/inventory_item/{sku}",
                    headers=headers,
                    json=inv_payload
                )
                print(f"[eBay] Inventory PUT for {sku}: status={inv_r.status_code}, body={inv_r.text[:500]}")

                # 2. Create offer
                modal = modal_items.get(str(listing["id"]), {}) if 'modal_items' in dir() else {}
                print(f"[eBay] Modal data for {listing['id']}: ship={modal.get('shipping_policy_id')}, ret={modal.get('return_policy_id')}, pay={modal.get('payment_policy_id')}")
                ship_pol = modal.get("shipping_policy_id")
                ret_pol = modal.get("return_policy_id")
                pay_pol = modal.get("payment_policy_id")
                listing_policies = {}
                if ship_pol: listing_policies["fulfillmentPolicyId"] = ship_pol
                if ret_pol: listing_policies["returnPolicyId"] = ret_pol
                if pay_pol: listing_policies["paymentPolicyId"] = pay_pol

                offer_payload = {
                    "sku": sku,
                    "marketplaceId": "EBAY_US",
                    "format": "FIXED_PRICE",
                    "listingDescription": description,
                    "pricingSummary": {
                        "price": {"value": str(price), "currency": "USD"}
                    },
                    "categoryId": category_id,
                    "merchantLocationKey": "default",
                    "listingPolicies": listing_policies
                }

                offer_r = _req3.post(
                    f"{api_base}/sell/inventory/v1/offer",
                    headers=headers,
                    json=offer_payload
                )
                print(f"[eBay] Offer POST status={offer_r.status_code}, body={offer_r.text[:500]}")
                offer_data = offer_r.json() if offer_r.text else {}
                offer_id = offer_data.get("offerId")

                # If offer already exists, find it and update
                if not offer_id and offer_r.status_code == 400:
                    errors = offer_data.get("errors", [])
                    for err in errors:
                        if err.get("errorId") == 25002:
                            for param in err.get("parameters", []):
                                if param.get("name") == "offerId":
                                    offer_id = param.get("value")
                                    print(f"[eBay] Reusing existing offer {offer_id}")
                                    # Update existing offer
                                    upd_r = _req3.put(
                                        f"{api_base}/sell/inventory/v1/offer/{offer_id}",
                                        headers=headers,
                                        json=offer_payload
                                    )
                                    print(f"[eBay] Offer UPDATE status={upd_r.status_code}")
                                    break

                if offer_id:
                    # Publish as scheduled listing 1 year out — appears in Seller Hub as scheduled
                    from datetime import datetime as _dt, timedelta as _td
                    scheduled_date = (_dt.utcnow() + _td(days=365)).strftime("%Y-%m-%dT%H:%M:%S.000Z")
                    listing_id = offer_id
                    try:
                        pub_r = _req3.post(
                            f"{api_base}/sell/inventory/v1/offer/{offer_id}/publish",
                            headers=headers,
                            json={"scheduledStartDate": scheduled_date}
                        )
                        print(f"[eBay] Scheduled publish for {offer_id}: status={pub_r.status_code}, body={pub_r.text[:300]}")
                        pub_data = pub_r.json() if pub_r.text else {}
                        listing_id = pub_data.get("listingId") or offer_id
                    except Exception as _pub_err:
                        print(f"[eBay] Publish warning: {_pub_err}")

                    try:
                        supabase.table("listings").update({
                            "ebay_item_id": listing_id,
                            "status": "ebay_scheduled"
                        }).eq("id", listing["id"]).execute()
                    except Exception as _db_err:
                        print(f"[eBay] DB update warning: {_db_err}")
                    results.append({"id": listing["id"], "ok": True, "offer_id": offer_id, "listing_id": listing_id})
                else:
                    results.append({"id": listing["id"], "ok": False, "error": str(offer_data)})
            except Exception as item_err:
                results.append({"id": listing.get("id","?"), "ok": False, "error": str(item_err)})

        ok_count = sum(1 for r in results if r["ok"])
        print(f"[eBay] Final: submitted={ok_count}/{len(results)}, results={results}")
        return {"ok": True, "submitted": ok_count, "total": len(results), "results": results, "draft": True, "message": "Listings created as drafts in eBay Seller Hub — review and publish there"}

    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))



@app.get("/api/ebay/whoami")
async def ebay_whoami(request: Request):
    """Check which eBay user account is connected."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        import requests as _req
        token = get_ebay_token(business_id)
        api_base = "https://api.ebay.com" if EBAY_ENV != "sandbox" else "https://api.sandbox.ebay.com"
        headers = {"Authorization": f"Bearer {token}"}
        # Check fulfillment policies to see seller info
        r = _req.get(f"{api_base}/sell/account/v1/fulfillment_policy?marketplace_id=EBAY_US", headers=headers)
        # Also check inventory items count
        r2 = _req.get(f"{api_base}/sell/inventory/v1/inventory_item?limit=1", headers=headers)
        inv_data = r2.json() if r2.ok else {}
        return {
            "policies_status": r.status_code,
            "policies_count": len(r.json().get("fulfillmentPolicies", [])) if r.ok else 0,
            "inventory_status": r2.status_code,
            "inventory_total": inv_data.get("total", 0),
            "inventory_items": [{"sku": i.get("sku"), "title": i.get("product",{}).get("title","")} for i in inv_data.get("inventoryItems", [])[:5]]
        }
    except Exception as e:
        return {"error": str(e)}


@app.post("/api/ebay/setup-location")
async def setup_ebay_location(request: Request):
    """Create a default merchant location on eBay (required for listings)."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        import requests as _req
        token = get_ebay_token(business_id)
        api_base = "https://api.ebay.com" if EBAY_ENV != "sandbox" else "https://api.sandbox.ebay.com"
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json", "Content-Language": "en-US"}

        location_payload = {
            "location": {
                "address": {
                    "addressLine1": body.get("address", "123 Main St"),
                    "city": body.get("city", "San Francisco"),
                    "stateOrProvince": body.get("state", "CA"),
                    "postalCode": body.get("zip", "94102"),
                    "country": "US"
                }
            },
            "locationInstructions": "Items ship from this location",
            "name": body.get("name", "Default Warehouse"),
            "merchantLocationStatus": "ENABLED",
            "locationTypes": ["WAREHOUSE"]
        }

        r = _req.post(
            f"{api_base}/sell/inventory/v1/location/default",
            headers=headers,
            json=location_payload
        )
        return {"status": r.status_code, "body": r.text[:500]}
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/ebay/locations")
async def get_ebay_locations(request: Request):
    """List existing merchant locations."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        import requests as _req
        token = get_ebay_token(business_id)
        api_base = "https://api.ebay.com" if EBAY_ENV != "sandbox" else "https://api.sandbox.ebay.com"
        headers = {"Authorization": f"Bearer {token}"}
        r = _req.get(f"{api_base}/sell/inventory/v1/location", headers=headers)
        return r.json() if r.ok else {"status": r.status_code, "body": r.text[:500]}
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/ebay/policies")
async def get_ebay_policies(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        import requests as _req
        token = get_ebay_token(business_id)
        api_base = "https://api.ebay.com" if EBAY_ENV != "sandbox" else "https://api.sandbox.ebay.com"
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        fp = _req.get(f"{api_base}/sell/account/v1/fulfillment_policy?marketplace_id=EBAY_US", headers=headers)
        fp_data = fp.json() if fp.ok else {}
        fulfillment = []
        for p in (fp_data.get("fulfillmentPolicies") or []):
            fulfillment.append({"id": p.get("fulfillmentPolicyId"), "name": p.get("name"), "default": False})
        if fulfillment: fulfillment[0]["default"] = True
        rp = _req.get(f"{api_base}/sell/account/v1/return_policy?marketplace_id=EBAY_US", headers=headers)
        rp_data = rp.json() if rp.ok else {}
        returns = []
        for p in (rp_data.get("returnPolicies") or []):
            returns.append({"id": p.get("returnPolicyId"), "name": p.get("name"), "default": False})
        if returns: returns[0]["default"] = True
        pp = _req.get(f"{api_base}/sell/account/v1/payment_policy?marketplace_id=EBAY_US", headers=headers)
        pp_data = pp.json() if pp.ok else {}
        payments = []
        for p in (pp_data.get("paymentPolicies") or []):
            payments.append({"id": p.get("paymentPolicyId"), "name": p.get("name"), "default": False})
        if payments: payments[0]["default"] = True
        return {"fulfillment": fulfillment, "returns": returns, "payments": payments}
    except Exception as e:
        return {"fulfillment": [], "returns": [], "payments": [], "error": str(e)}

@app.get("/api/export/ebay-csv")
async def export_ebay_csv():
    import csv, io
    from fastapi.responses import StreamingResponse
    from datetime import datetime
    try:
        res = supabase.table("listings").select(
            "title,description,price,price_used,price_new,quantity,condition,photo_id,ebay_category_id,description"
        ).neq("status", "archived").execute()
        items = res.data or []
    except Exception as e:
        raise HTTPException(500, str(e))

    output = io.StringIO()

    # eBay draft flat file headers — same format as the working version
    output.write('#INFO,Version=0.0.2,Template= eBay-draft-listings-template_US,,,,,,,,\n')
    output.write('#INFO Action and Category ID are required fields.,,,,,,,,,,\n')
    output.write('#INFO,,,,,,,,,,\n')
    output.write('Action(SiteID=US|Country=US|Currency=USD|Version=1193|CC=UTF-8),Custom label (SKU),Category ID,Title,UPC,Price,Quantity,Item photo URL,Condition ID,Description,Format\n')

    writer = csv.writer(output, quoting=csv.QUOTE_ALL)

    for item in items:
        cond = str(item.get("condition") or "used").strip().lower()
        cond_id = "NEW" if cond == "new" else "USED"
        pid = str(item.get("photo_id") or "")
        # Look up all photos for this listing via group_photos table
        try:
            _gp = supabase.table("group_photos").select("photo_id").eq("group_id",
                (supabase.table("group_photos").select("group_id").eq("photo_id", pid).execute().data or [{}])[0].get("group_id", "")
            ).execute()
            _all_pids = [r["photo_id"] for r in (_gp.data or [])] if _gp.data else [pid]
            pic = "|".join(photo_url(p) for p in _all_pids if p) if _all_pids else (photo_url(pid) if pid else "")
        except Exception:
            pic = photo_url(pid) if pid else ""
        category_id = "12576"
        price = float(item.get("price") or item.get("price_used") or 0)
        writer.writerow([
            "Draft",
            "",
            category_id,
            str(item.get("title",""))[:80],
            "",
            f"{price:.2f}",
            str(int(item.get("quantity") or 1)),
            pic,
            cond_id,
            str(item.get('description','') or EBAY_DESCRIPTION),
            "FixedPrice",
        ])

    csv_bytes = output.getvalue().encode("utf-8")
    fn = f"ebay_export_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fn}"}
    )


@app.get("/api/photos/view/{photo_id}")
async def view_photo(photo_id: str, t: str = ""):
    from fastapi.responses import Response
    img_bytes = supabase.storage.from_("part-photos").download(photo_id)
    return Response(content=img_bytes, media_type="image/jpeg", headers={
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
        "Expires": "0"
    })

@app.post("/api/photos/rotate")
async def rotate_photo(request: Request):
    from PIL import Image
    import io
    body = await request.json()
    photo_id = body.get("photo_id", "")
    if not photo_id:
        raise HTTPException(400, "photo_id required")
    try:
        img_bytes = supabase.storage.from_("part-photos").download(photo_id)
        img = Image.open(io.BytesIO(img_bytes))
        direction = body.get('direction', 'cw')
        img = img.rotate(90 if direction == 'ccw' else -90, expand=True)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=90)
        buf.seek(0)
        supabase.storage.from_("part-photos").upload(
            path=photo_id,
            file=buf.read(),
            file_options={"content-type": "image/jpeg", "upsert": "true"}
        )
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/reset-queue")
async def reset_queue():
    try:
        supabase.table("listing_groups").update({"status": "waiting"}).in_("status", ["processing", "pending"]).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/stats")
async def get_stats(request: Request):
    try:
        business_id = require_auth(request)
        if not business_id:
            return {"total": 0, "processing": 0, "value": 0}
        res = supabase.table("listings").select("price, status").eq("business_id", business_id).neq("status", "archived").execute()
        items = res.data or []
        total = len(items)
        value = sum(float(i.get("price") or 0) for i in items)
        pending = supabase.table("listing_groups").select("id").eq("business_id", business_id).in_("status", ["pending", "processing"]).execute()
        processing = len(pending.data or [])
        return {"total": total, "processing": processing, "value": round(value, 2)}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.patch("/api/listings/{item_id}")
async def update_listing(item_id: str, body: UpdateField):
    try:
        supabase.table("listings").update({body.field: body.value}).eq("id", item_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/listings/{item_id}/rescan")
async def rescan_listing(item_id: str):
    try:
        res = supabase.table("listings").select("photo_id").eq("id", item_id).limit(1).execute()
        pid = (res.data[0].get("photo_id", "") if res.data else "")
        if pid:
            grp = supabase.table("group_photos").select("group_id").eq("photo_id", pid).limit(1).execute()
            if grp.data:
                gid = grp.data[0]["group_id"]
                supabase.table("listing_groups").update({"status": "pending"}).eq("id", gid).execute()
        supabase.table("listings").update({"status": "pending", "title": "Scanning..."}).eq("id", item_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/listings/archive-batch")
async def archive_batch():
    try:
        res = supabase.table("listings").select("*").neq("status", "archived").execute()
        items = res.data or []
        if items:
            ids = [str(i["id"]) for i in items]
            supabase.table("listings").update({"status": "archived"}).in_("id", ids).execute()
        return {"ok": True, "count": len(items)}
    except Exception as e:
        raise HTTPException(500, str(e))

# ── API: BATCH UPLOAD ─────────────────────────────────────────── #

class CreateGroup(BaseModel):
    session_id: str
    condition:  str


# ── API: SAVED BATCHES ────────────────────────────────────────── #

class SaveBatchBody(BaseModel):
    name: str

@app.post("/api/saved-batches")
async def create_saved_batch(request: Request, body: SaveBatchBody):
    """Snapshot all current (non-archived) listings into a new named folder."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(400, "Folder name required")
    try:
        res = supabase.table("listings")\
            .select("*")\
            .eq("business_id", business_id)\
            .neq("status", "archived")\
            .order("created_at", desc=True)\
            .execute()
        listings = res.data or []

        primary_pids = [str(l.get("photo_id") or "") for l in listings if l.get("photo_id")]
        group_photo_map = {}
        if primary_pids:
            try:
                gp_res = supabase.table("group_photos").select("group_id, photo_id").in_("photo_id", primary_pids[:100]).execute()
                pid_to_gid = {row["photo_id"]: row["group_id"] for row in (gp_res.data or [])}
                group_ids = list(set(pid_to_gid.values()))
                if group_ids:
                    all_gp = supabase.table("group_photos").select("group_id, photo_id").in_("group_id", group_ids).execute()
                    gid_to_photos = {}
                    for row in (all_gp.data or []):
                        gid_to_photos.setdefault(row["group_id"], []).append(row["photo_id"])
                    for pid, gid in pid_to_gid.items():
                        group_photo_map[pid] = gid_to_photos.get(gid, [pid])
            except Exception as gp_err:
                print(f"   saved-batch photo group lookup failed: {gp_err}")

        folder_res = supabase.table("saved_batches").insert({
            "business_id": business_id,
            "name": name,
        }).execute()
        if not folder_res.data:
            raise HTTPException(500, "Failed to create folder")
        folder_id = folder_res.data[0]["id"]

        snapshot_rows = []
        for l in listings:
            pid = str(l.get("photo_id") or "")
            all_photos = group_photo_map.get(pid, [pid] if pid else [])
            snapshot_rows.append({
                "saved_batch_id":      folder_id,
                "business_id":         business_id,
                "original_listing_id": l.get("id"),
                "title":               l.get("title"),
                "description":         l.get("description"),
                "price":               l.get("price"),
                "price_used":          l.get("price_used"),
                "price_new":           l.get("price_new"),
                "quantity":            l.get("quantity"),
                "condition":           l.get("condition"),
                "listing_type":        l.get("listing_type"),
                "ebay_category_id":    l.get("ebay_category_id"),
                "status":              l.get("status"),
                "photo_id":            pid or None,
                "photo_ids":           all_photos,
                "raw":                 l,
            })

        if snapshot_rows:
            for i in range(0, len(snapshot_rows), 100):
                supabase.table("saved_batch_listings").insert(snapshot_rows[i:i+100]).execute()

        return JSONResponse({
            "id": folder_id,
            "name": name,
            "count": len(snapshot_rows),
        })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/saved-batches")
async def list_saved_batches(request: Request):
    """Return all saved folders for the current business with item counts."""
    business_id = require_auth(request)
    if not business_id:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    try:
        folders_res = supabase.table("saved_batches")\
            .select("id, name, created_at")\
            .eq("business_id", business_id)\
            .order("created_at", desc=True)\
            .execute()
        folders = folders_res.data or []

        folder_ids = [f["id"] for f in folders]
        counts = {}
        if folder_ids:
            cnt_res = supabase.table("saved_batch_listings")\
                .select("saved_batch_id")\
                .in_("saved_batch_id", folder_ids)\
                .execute()
            for row in (cnt_res.data or []):
                bid = row["saved_batch_id"]
                counts[bid] = counts.get(bid, 0) + 1

        for f in folders:
            f["count"] = counts.get(f["id"], 0)

        return JSONResponse(folders)
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/saved-batches/{batch_id}")
async def get_saved_batch(batch_id: str, request: Request):
    """Return all snapshot listings for a folder, scoped to current business."""
    business_id = require_auth(request)
    if not business_id:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    try:
        folder_res = supabase.table("saved_batches")\
            .select("id, name, created_at")\
            .eq("id", batch_id)\
            .eq("business_id", business_id)\
            .limit(1)\
            .execute()
        if not folder_res.data:
            raise HTTPException(404, "Folder not found")
        folder = folder_res.data[0]

        items_res = supabase.table("saved_batch_listings")\
            .select("*")\
            .eq("saved_batch_id", batch_id)\
            .eq("business_id", business_id)\
            .order("id", desc=False)\
            .execute()
        items = items_res.data or []

        for it in items:
            pid = str(it.get("photo_id") or "")
            all_pids = it.get("photo_ids") or ([pid] if pid else [])
            it["thumb_url"]  = photo_url(pid, thumb=True)
            it["full_url"]   = photo_url(pid)
            it["all_photos"] = [{"thumb": photo_url(p, thumb=True), "full": photo_url(p)} for p in all_pids if p]
            it["price"]      = float(it.get("price") or 0)
            it["price_used"] = float(it.get("price_used") or 0)
            it["price_new"]  = float(it.get("price_new") or 0)
            it["quantity"]   = int(it.get("quantity") or 1)

        return JSONResponse({
            "id":         folder["id"],
            "name":       folder["name"],
            "created_at": folder["created_at"],
            "items":      items,
        })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/api/saved-batches/{batch_id}")
async def delete_saved_batch(batch_id: str, request: Request):
    """Delete a folder and all its snapshot rows (cascade handles the rows)."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        chk = supabase.table("saved_batches")\
            .select("id")\
            .eq("id", batch_id)\
            .eq("business_id", business_id)\
            .limit(1)\
            .execute()
        if not chk.data:
            raise HTTPException(404, "Folder not found")

        supabase.table("saved_batches").delete().eq("id", batch_id).eq("business_id", business_id).execute()
        return JSONResponse({"ok": True})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/groups/pending")
async def get_pending_groups(request: Request):
    business_id = require_auth(request)
    if not business_id:
        from fastapi.responses import JSONResponse
        return JSONResponse({"groups": []}, status_code=401)
    try:
        groups = supabase.table("listing_groups").select("id,status,condition,created_at").eq("business_id", business_id).in_("status", ["waiting", "pending", "processing"]).order("created_at", desc=True).execute()
        result = []
        for g in (groups.data or []):
            photos = supabase.table("group_photos").select("photo_id").eq("group_id", g["id"]).execute()
            photo_ids = [p["photo_id"] for p in (photos.data or [])]
            if not photo_ids:
                continue
            result.append({
                "id": g["id"],
                "status": g["status"],
                "condition": g["condition"],
                "created_at": g["created_at"],
                "photo_id": photo_ids[0],
                "photo_count": len(photo_ids),
                "photo_url": f"{SUPABASE_URL}/storage/v1/object/public/part-photos/{photo_ids[0]}"
            })
        return {"groups": result}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/groups")
async def create_group(body: CreateGroup, request: Request):
    try:
        business_id = require_auth(request)
        # Check scan limit
        biz = supabase.table("businesses").select("scan_count,scan_limit").eq("id", business_id).execute()
        if biz.data:
            scan_count = biz.data[0].get("scan_count") or 0
            scan_limit = biz.data[0].get("scan_limit") or 50
            if scan_count >= scan_limit:
                raise HTTPException(402, "Scan limit reached. Please contact us to upgrade your plan.")
            # Increment scan count
            supabase.table("businesses").update({"scan_count": scan_count + 1}).eq("id", business_id).execute()
        res = supabase.table("listing_groups").insert({
            "session_id": body.session_id,
            "status":     "waiting",
            "quantity":   1,
            "condition":  body.condition,
            "business_id": business_id,
        }).execute()
        import traceback
        print(f"Group insert result: {res}")
        data = res.data
        gid = data[0]["id"] if isinstance(data, list) and data else (data.get("id") if isinstance(data, dict) else None)
        if not gid:
            raise Exception(f"No ID returned. data={data}")
        return {"id": gid}
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))

class SubmitGroup(BaseModel):
    group_id:  str
    condition: str
    quantity:  int

@app.post("/api/groups/submit")
async def submit_group(body: SubmitGroup):
    try:
        supabase.table("listing_groups").update({
            "condition": body.condition,
            "quantity":  body.quantity,
            "status":    "pending",
        }).eq("id", body.group_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/photos/upload")
async def upload_photo(request: Request):
    try:
        form     = await request.form()
        file     = form["file"]
        gid      = str(form["group_id"])
        idx      = int(form.get("index", 0))
        contents = await file.read()
        dt  = datetime.now()
        fn  = f"{dt.strftime('%d%m%y')}_{dt.strftime('%H%M%S')}_{idx}.jpg"
        print(f"Uploading photo: {fn}, size={len(contents)}, group={gid}")
        supabase.storage.from_("part-photos").upload(
            path=fn,
            file=contents,
            file_options={"content-type": "image/jpeg", "upsert": "true"}
        )
        supabase.table("group_photos").insert({"group_id": gid, "photo_id": fn}).execute()
        return {"ok": True, "photo_id": fn, "url": photo_url(fn, thumb=True)}
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))



# ── API: FULL DEEP RESEARCH ──────────────────────────────────── #

@app.post("/api/auction/deep-research-full")
async def deep_research_full(request: Request):
    import os, json, base64, asyncio, fitz
    from concurrent.futures import ThreadPoolExecutor
    import google.generativeai as genai

    form = await request.form()
    items_json = form.get("items", "[]")
    items = json.loads(items_json)
    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(400, "GEMINI_API_KEY not set")

    pdf_bytes = None
    pdf_file = form.get("pdf")
    if pdf_file and hasattr(pdf_file, "read"):
        pdf_bytes = await pdf_file.read()

    genai.configure(api_key=gemini_key)
    model = genai.GenerativeModel("gemini-2.5-flash")
    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)

    def extract_single_image(pdf_bytes, img_index):
        try:
            import fitz
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            seen = set()
            all_images = []
            for page in doc:
                for img in page.get_images(full=True):
                    xref = img[0]
                    if xref in seen: continue
                    seen.add(xref)
                    bi = doc.extract_image(xref)
                    if bi and len(bi.get("image","")) > 8000:
                        all_images.append(bi["image"])
            doc.close()
            if img_index < len(all_images):
                return [all_images[img_index]]
            return []
        except Exception as e:
            print(f"extract_single_image error: {e}")
            return []

    def extract_page_image(pdf_bytes, page_start, page_end):
        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            images = []
            for page_num in range(page_start - 1, min(page_end, len(doc))):
                page = doc[page_num]
                mat = fitz.Matrix(2.0, 2.0)
                pix = page.get_pixmap(matrix=mat)
                images.append(pix.tobytes("jpeg"))
            doc.close()
            return images
        except Exception as e:
            print(f"Image extract error: {e}")
            return []

    def identify_item_from_image(images, title):
        if not images:
            return title
        try:
            id_prompt = f"You are an auction appraiser. Use BOTH the image AND the listing title equally to identify this item. Title: {title}. Look at the image for: exact model numbers on labels/nameplates, brand logos, condition, visible accessories. Combine both sources and return one precise description. Do not ignore the title — it may contain info not visible in the image."
            parts = [id_prompt] + [{"mime_type": "image/jpeg", "data": img} for img in images[:1]]
            r = model.generate_content(parts, generation_config={"max_output_tokens": 150})
            result = r.text.strip().strip('"')
            return result if result else title
        except Exception as e:
            print(f"Image ID error: {e}")
            return title

    def clean_title(raw_title):
        """Strip address fragments, company boilerplate, and catalog noise from auction titles."""
        import re
        t = raw_title
        # Remove street addresses like "Siemensstrasse 7", "123 Main St"
        t = re.sub(r'\d+\s+[A-Z][a-z]+(?:strasse|street|ave|blvd|rd|st|dr|ln|way)', '', t, flags=re.IGNORECASE)
        t = re.sub(r'[A-Z][a-z]+(?:strasse|gasse|platz|weg)\s+\d+', '', t, flags=re.IGNORECASE)
        # Remove street addresses
        t = re.sub(r'\\b[A-Za-z]+(?:strasse|gasse|weg|strase)\\b\\s*\\d*', '', t, flags=re.IGNORECASE)
        # Remove "GmbH", "Inc", "LLC", "Ltd", "Corp", "Co." standalone
        t = re.sub(r'\b(?:GmbH|Inc\.?|LLC|Ltd\.?|Corp\.?|Co\.)\b', '', t, flags=re.IGNORECASE)
        # Remove loading fee notes
        t = re.sub(r'Loading Fee[:\s]*\$?\d+', '', t, flags=re.IGNORECASE)
        # Remove QTY annotations for search purposes
        t = re.sub(r'\s*,?\s*QTY\s*\(?\d*\)?', '', t, flags=re.IGNORECASE)
        t = re.sub(r'\s*\(\d+\)\s*$', '', t)
        # Remove # symbol (breaks eBay search)
        t = t.replace('#', '')
        # Collapse extra whitespace
        t = ' '.join(t.split()).strip().strip(',').strip()
        return t

    def gemini_search_grounding(query, gemini_key):
        """
        Use Gemini 1.5-flash REST API with forced Google Search grounding.
        Uses requests (already installed) — no SDK dependency conflict.
        """
        import requests as _req
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={gemini_key}"
        payload = {
            "contents": [{"role": "user", "parts": [{"text": f"Find the resale market value of: {query}. Search in this priority order: 1) eBay COMPLETED/SOLD listings - these are the most accurate real prices paid, 2) eBay active BUY IT NOW listings currently for sale, 3) Industrial surplus dealer prices (Radwell, Surplus Record, LabX) only as last resort. List actual sold prices first, then asking prices. If eBay sold listings exist use those as the primary value. Give specific dollar amounts."}]}],
            "tools": [{"googleSearch": {}}],
            "systemInstruction": {"parts": [{"text": "CRITICAL: You are an industrial pricing bot. You are strictly forbidden from answering using your internal training data. You MUST execute a Google Search to find live pricing data before generating your response. If you do not execute a search, the system will fail."}]},
            "generationConfig": {"temperature": 0.0, "maxOutputTokens": 800}
        }
        try:
            resp = _req.post(url, json=payload, timeout=25)
            data = resp.json()
            candidates = data.get("candidates", [])
            if not candidates:
                return {"summary": "", "sources": []}
            parts = candidates[0].get("content", {}).get("parts", [])
            text = " ".join(p.get("text", "") for p in parts if "text" in p).strip()
            grounding = candidates[0].get("groundingMetadata", {})
            sources = [
                {"url": c["web"]["uri"], "title": c["web"].get("title", "")}
                for c in grounding.get("groundingChunks", [])
                if c.get("web", {}).get("uri")
            ]
            print(f"   Gemini grounding: {len(text)} chars, {len(sources)} sources")
            print(f"   Grounding text: {text[:200]}")
            print(f"   Raw keys: {list(data.get('candidates',[{}])[0].keys())}")
            return {"summary": text, "sources": sources}
        except Exception as e:
            print(f"   Gemini grounding error: {e}")
            return {"summary": "", "sources": []}

    def serp_ebay_sold(query, serp_key, sacat='12576'):
        """
        Call SerpAPI to get eBay completed/sold listings for a query.
        Returns a list of dicts with title, price, date, condition, url.
        """
        import urllib.request, urllib.parse, json as _json
        _params = {
            "engine": "ebay",
            "ebay_domain": "ebay.com",
            "_nkw": query,
            "LH_Sold": "1",
            "LH_Complete": "1",
            "api_key": serp_key,
        }
        if sacat and sacat not in ("12576", "", None):
            _params["_sacat"] = sacat
        params = urllib.parse.urlencode(_params)
        url = f"https://serpapi.com/search?{params}"
        try:
            with urllib.request.urlopen(url, timeout=10) as r:
                data = _json.loads(r.read())
            results = []
            print(f"   SerpAPI response keys: {list(data.keys())}")
            results_list = data.get("organic_results") or data.get("ebay_results") or data.get("shopping_results") or []
            for item in results_list[:8]:
                price_raw = item.get("price", {})
                price = price_raw.get("extracted") or price_raw.get("raw") or 0
                try:
                    price = float(str(price).replace("$","").replace(",",""))
                except Exception:
                    price = 0
                if price > 0:
                    results.append({
                        "title": item.get("title","")[:80],
                        "price": price,
                        "condition": item.get("condition","Used"),
                        "date": item.get("selling_states",{}).get("sold_date","") or "",
                        "url": item.get("link",""),
                    })
            return results
        except Exception as e:
            print(f"   SerpAPI error: {e}")
            return []

    def research_item(item, images):
        import os as _os
        title = item.get("title", "")
        lot = item.get("lot", "")
        current_val = item.get("your_value", 0) or 0
        serp_key = _os.getenv("SERP_API_KEY", "")

        # Clean title before research — remove address/company junk
        clean = clean_title(title)
        if clean != title:
            print(f"Lot {lot} title cleaned: '{title}' → '{clean}'")

        # Step 1: identify exact model from image
        identified = identify_item_from_image(images, clean)
        if identified != clean:
            print(f"Lot {lot} image ID: {identified}")

        # Step 2: Gemini Search Grounding for real market pricing
        serp_results = []
        serp_context = ""
        _gsummary = ""
        if gemini_key:
            _grounding = gemini_search_grounding(clean, gemini_key)
            _gsummary = _grounding.get("summary", "")
            if _gsummary:
                serp_context = f"""MARKET RESEARCH DATA (from live Google Search — use as PRIMARY pricing source):
{_gsummary}

Extract specific dollar amounts from the above. Base revised_value on actual prices found.
Do NOT ignore this data. Do NOT use your training knowledge if this data contradicts it.
"""
        # Legacy SerpAPI block (disabled — kept for fallback reference)
        if False and serp_key:
            # --- Pre-classification: get eBay _sacat and negative keywords ---
            _sacat_map = {
                "12576": "Business & Industrial - Other",
                "58058": "Lasers & Laser Optics (industrial/scientific lasers, fiber lasers, CO2 lasers, laser systems)",
                "105595": "Laser Accessories & Parts",
                "11804": "CNC, Metalworking & Manufacturing",
                "11808": "Electrical Equipment & Supplies",
                "11803": "Semiconductor & PCB Equipment",
                "78989": "Test, Measurement & Inspection",
                "4666":  "Pumps & Plumbing",
                "11816": "Hydraulics, Pneumatics & Plumbing",
                "11815": "Healthcare, Lab & Dental",
                "3673":  "Computers & Networking",
                "58058": "Lasers & Laser Accessories",
                "11700": "Consumer Electronics",
                "26230": "Hand Tools",
                "92074": "Power Tools",
            }
            _cat_prompt = f"""You are an eBay category classifier for industrial equipment.

Item: {clean}

Choose the single best eBay category ID from this list:
{chr(10).join(f'  {k}: {v}' for k,v in _sacat_map.items())}

Also decide if negative keywords are needed to filter out medical/consumer results.
Negative keywords to consider: -medical -dental -cosmetic -hair -aesthetic -salon

Respond ONLY with valid JSON, no markdown:
{{"sacat": "12576", "negative_keywords": "-medical -dental", "is_industrial": true}}

If unsure about negative keywords, use empty string for negative_keywords."""

            _sacat = "12576"
            _negative_kw = ""
            _is_industrial = True
            try:
                _cat_response = model.generate_content(
                    _cat_prompt,
                    generation_config={"max_output_tokens": 100, "temperature": 0}
                )
                _cat_text = _cat_response.text.strip()
                if "```" in _cat_text:
                    _cat_text = _cat_text.split("```")[1]
                    if _cat_text.startswith("json"):
                        _cat_text = _cat_text[4:]
                _cat_text = _cat_text.strip()
                # Find the JSON object
                _s = _cat_text.find("{")
                _e = _cat_text.rfind("}") + 1
                if _s >= 0 and _e > _s:
                    _cat_text = _cat_text[_s:_e]
                import json as _json2
                from json_repair import repair_json as _rj
                _cat_data = _json2.loads(_rj(_cat_text))
                _sacat = str(_cat_data.get("sacat", "12576"))
                _negative_kw = str(_cat_data.get("negative_keywords", ""))
                _is_industrial = bool(_cat_data.get("is_industrial", True))
                print(f"   Category: {_sacat_map.get(_sacat, _sacat)}, industrial={_is_industrial}, negatives='{_negative_kw}'")
            except Exception as _ce:
                print(f"   Category pre-classification failed: {_ce}, using default sacat=12576")

            # Build search query with phrase matching + negative keywords
            _w = clean.split()
            _base_query = '"' + clean + '"' if len(_w) >= 3 else clean
            search_query = (_base_query + " " + _negative_kw).strip()
            print(f"   SerpAPI eBay sold search: '{search_query}' (sacat={_sacat})")
            serp_results = serp_ebay_sold(search_query, serp_key, sacat=_sacat)

            # --- IQR variance-based sanity check ---
            if serp_results:
                prices = sorted([r["price"] for r in serp_results])
                n = len(prices)
                if n >= 4:
                    q1 = prices[n // 4]
                    q3 = prices[(3 * n) // 4]
                    iqr = q3 - q1
                    median = prices[n // 2]
                    cv = (iqr / median) if median > 0 else 1
                    if cv > 1.5:
                        print(f"   SerpAPI results discarded — high variance (CV={cv:.2f}), mixed categories likely")
                        serp_results = []
                    else:
                        print(f"   SerpAPI variance OK: IQR=${iqr:.0f}, CV={cv:.2f}, median=${median:.0f}")
                elif n >= 2:
                    # Small sample: check if range is >5x spread
                    _spread = prices[-1] / prices[0] if prices[0] > 0 else 10
                    if _spread > 5:
                        print(f"   SerpAPI results discarded — spread too wide ({prices[0]:.0f}-{prices[-1]:.0f})")
                        serp_results = []

            if serp_results:
                prices = [r["price"] for r in serp_results]
                avg = sum(prices) / len(prices)
                low = min(prices)
                high = max(prices)
                lines = [f"  - ${r['price']:.0f} — {r['title']} ({r['condition']}) {r['date']}" for r in serp_results]
                serp_context = f"""
REAL EBAY SOLD DATA (from live eBay completed listings — use this as primary pricing source):
Found {len(serp_results)} sold comps: low ${low:.0f}, high ${high:.0f}, avg ${avg:.0f}
{chr(10).join(lines)}

Base your revised_value on these actual sold prices. Do not override this with guesses.
"""
                print(f"   SerpAPI: {len(serp_results)} comps, avg ${avg:.0f}, range ${low:.0f}-${high:.0f}")
            else:
                serp_context = "No eBay sold comps found via SerpAPI - use web search grounding for pricing."
                print(f"   SerpAPI: no results for '{search_query}'")

        prompt = f"""You are an expert industrial machinery appraiser and secondary market researcher.
Your job is to determine the actual cash value of an industrial asset at auction.

--- ITEM DETAILS ---
Lot: #{lot}
Clean Title: {clean}
Image-Identified Model: {identified}
Initial Estimate: ${current_val}

--- MARKET RESEARCH DATA ---
{serp_context}

--- PRICING HIERARCHY RULES (CRITICAL) ---
You must evaluate the MARKET RESEARCH DATA using this strict waterfall hierarchy. Do NOT skip tiers.

TIER 1: SOLD/COMPLETED LISTINGS (Highest Priority)
If the data contains actual verified sold prices, base your estimate entirely on these. Ignore all asking prices.
-> pricing_tier = "SOLD_COMPS"

TIER 2: ACTIVE MARKETPLACE LISTINGS (The Ceiling)
If no sold data exists, look for active listings on open marketplaces (eBay, etc).
Rule: The LOWEST reasonable active listing establishes the absolute CEILING of value. A buyer will not pay $8,000 if they can buy it right now on eBay for $3,995.
Calculation: Find the lowest active price. Apply a 15-25% discount to estimate actual sell price. Ignore high-priced outliers.
-> pricing_tier = "ASKING_PRICES"

TIER 3: INDUSTRIAL DEALER ASKING PRICES (Last Resort Anchor)
If NO marketplace data exists, use retail/surplus dealer asking prices (Radwell, PLC Center, etc).
Rule: Dealers charge massive premiums. Apply a 40-60% discount to find auction/resale cash value.
-> pricing_tier = "ASKING_PRICES"

TIER 4: NO DATA
If the MARKET RESEARCH DATA contains no dollar values relevant to this item, admit it.
-> pricing_tier = "NO_DATA"

--- HALLUCINATION GUARDRAILS ---
- You are FORBIDDEN from using pricing_tier "SOLD_COMPS" unless the word "sold" or "completed" is explicitly in the data.
- Do NOT average a $3,995 eBay listing with a $15,000 dealer listing. The $3,995 becomes the absolute ceiling.
- Do NOT fabricate comps. Only list prices explicitly found in the MARKET RESEARCH DATA above.
- confidence must be "high" only with 3+ verified sold comps, otherwise "medium" or "low".

SHIPPING WEIGHT: Estimate from item type and visible size.

Return ONLY valid JSON, no markdown:
{{"revised_value": 3200, "confidence": "medium", "pricing_tier": "ASKING_PRICES", "pricing_flag": "Based on lowest active eBay listing $3,995 minus 20% discount", "comps": [{{"title": "Item name", "price": 3995, "date": "Apr 2025", "source": "eBay Active"}}], "image_notes": "What the image shows", "recommendation": "watch", "rec_reason": "One active eBay listing at $3,995 sets ceiling, estimated sell price $3,200", "notes": "Market summary with sources", "weight_item_lbs": 50.0, "weight_packaged_lbs": 55.0, "weight_note": "Estimated", "liquidity_score": 2, "liquidity_note": "Limited market data", "sold_30d": 0, "sold_90d": 0, "active_listings": 1}}

pricing_tier values: SOLD_COMPS | ASKING_PRICES | MSRP_ONLY | COMPARABLE_ITEMS | NO_DATA
weight fields: use null if truly unknown"""

        # Use Gemini with search grounding if available
        try:
            from google import genai as _gc
            from google.genai import types as _gt
            _client = _gc.Client(api_key=gemini_key)
            _parts = [prompt]
            for img_bytes in images[:2]:
                _parts.append(_gt.Part.from_bytes(data=img_bytes, mime_type="image/jpeg"))
            _cfg = _gt.GenerateContentConfig(
                tools=[_gt.Tool(google_search=_gt.GoogleSearch())],
                max_output_tokens=1500
            )
            _resp = _client.models.generate_content(
                model="gemini-2.5-flash",
                contents=_parts,
                config=_cfg
            )
            response = _resp
            # Extract grounding from new SDK response
            try:
                gm = _resp.candidates[0].grounding_metadata
                if gm and gm.search_entry_point:
                    ai_overview_html = gm.search_entry_point.rendered_content or ""
                for chunk in (gm.grounding_chunks or []):
                    if hasattr(chunk, "web") and chunk.web:
                        grounding_sources.append({"title": chunk.web.title or "", "uri": chunk.web.uri or ""})
            except Exception:
                pass
            # Make .text work for downstream parsing
            class _Wrap:
                def __init__(self, r): self._r = r
                @property
                def text(self): return self._r.text
                @property
                def candidates(self): return self._r.candidates
            response = _Wrap(_resp)
        except Exception as search_err:
            print(f"   Search grounding failed: {search_err}")
            parts = [prompt]
            for img_bytes in images[:2]:
                parts.append({"mime_type": "image/jpeg", "data": img_bytes})
            response = model.generate_content(parts, generation_config={"max_output_tokens": 1500})
        # Extract AI overview + sources from grounding metadata
        ai_overview_html = ""
        grounding_sources = []
        try:
            candidates = response.candidates
            if candidates:
                gm = getattr(candidates[0], "grounding_metadata", None)
                if gm:
                    sep = getattr(gm, "search_entry_point", None)
                    if sep:
                        ai_overview_html = getattr(sep, "rendered_content", "") or ""
                    chunks = getattr(gm, "grounding_chunks", []) or []
                    for chunk in chunks:
                        web = getattr(chunk, "web", None)
                        if web:
                            grounding_sources.append({
                                "title": getattr(web, "title", ""),
                                "uri":   getattr(web, "uri", ""),
                            })
        except Exception as gm_err:
            print(f"   Grounding metadata error: {gm_err}")

        raw = response.text.strip()
        print(f"   Deep research raw response (lot {lot}): {raw[:2000]}")
        try:
            _d = json.loads(raw if raw.startswith("{") else raw[raw.find("{"):raw.rfind("}")+1])
            print(f"   notes: {_d.get(chr(110)+chr(111)+chr(116)+chr(101)+chr(115),chr(101)+chr(109)+chr(112)+chr(116)+chr(121))}")
        except: pass
        print(f"   AI overview chars: {len(ai_overview_html)}, sources: {len(grounding_sources)}")
        # Strip markdown fences
        if "```" in raw:
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip()
        raw = " ".join(raw.splitlines())

        raw = " ".join(raw.splitlines())
        # Find JSON object boundaries
        start = raw.find("{")
        end = raw.rfind("}") + 1
        if start >= 0 and end > start:
            raw = raw[start:end]
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            from json_repair import repair_json
            data = json.loads(repair_json(raw))
        # Sanitize string fields to prevent SSE encoding issues
        for key in ["image_notes", "rec_reason", "notes", "confidence", "recommendation", "pricing_tier", "pricing_flag", "liquidity_note", "weight_note"]:
            if key in data:
                data[key] = str(data[key]).replace("\n", " ").replace("\r", " ")
        if "comps" in data:
            for comp in data["comps"]:
                for k in comp:
                    comp[k] = str(comp[k]).replace("\n", " ") if isinstance(comp[k], str) else comp[k]
        data["ai_overview_html"] = ai_overview_html
        # Inject grounding summary into notes if available
        if _gsummary and not data.get("notes"):
            data["notes"] = _gsummary
        data["grounding_sources"] = grounding_sources
        # If both SerpAPI and grounding failed, override any hallucinated high confidence
        if not serp_results and not ai_overview_html and not _gsummary:
            if data.get("pricing_tier") == "SOLD_COMPS" and data.get("confidence") == "high":
                data["confidence"] = "low"
                data["pricing_tier"] = "NO_DATA"
                data["pricing_flag"] = "No verified data sources available - estimate may not reflect actual market"

        # --- Hybrid Escalation: call gemini-2.5-pro for hard items ---
        escalate_tiers = {"NO_DATA", "COMPARABLE_ITEMS", "MSRP_ONLY"}
        if data.get("pricing_tier") in escalate_tiers and gemini_key:
            print(f"   Escalating lot {lot} to gemini-2.5-pro (tier={data.get('pricing_tier')})")
            try:
                import requests as _req
                pro_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-pro:generateContent?key={gemini_key}"
                pro_payload = {
                    "contents": [{"role": "user", "parts": [{"text": prompt}]}],
                    "generationConfig": {"temperature": 0.0, "maxOutputTokens": 2000}
                }
                pro_resp = _req.post(pro_url, json=pro_payload, timeout=60)
                pro_data = pro_resp.json()
                pro_parts = pro_data.get("candidates", [{}])[0].get("content", {}).get("parts", [])
                pro_raw = " ".join(p.get("text", "") for p in pro_parts if "text" in p).strip()
                if pro_raw:
                    if "```" in pro_raw:
                        pro_raw = pro_raw.split("```")[1]
                        if pro_raw.startswith("json"): pro_raw = pro_raw[4:]
                    pro_raw = pro_raw.strip()
                    s = pro_raw.find("{"); e = pro_raw.rfind("}") + 1
                    if s >= 0 and e > s:
                        pro_raw = pro_raw[s:e]
                    try:
                        pro_result = json.loads(pro_raw)
                    except Exception:
                        from json_repair import repair_json
                        pro_result = json.loads(repair_json(pro_raw))
                    # Sanitize and merge
                    for key in ["image_notes","rec_reason","notes","confidence","recommendation","pricing_tier","pricing_flag","liquidity_note","weight_note"]:
                        if key in pro_result:
                            pro_result[key] = str(pro_result[key]).replace("\n"," ").replace("\r"," ")
                    pro_result["model_used"] = "gemini-2.5-pro"
                    pro_result["ai_overview_html"] = ai_overview_html
                    pro_result["grounding_sources"] = grounding_sources
                    if _gsummary and not pro_result.get("notes"):
                        pro_result["notes"] = _gsummary
                    print(f"   Pro escalation result: tier={pro_result.get('pricing_tier')}, value={pro_result.get('revised_value')}")
                    return pro_result
            except Exception as _pe:
                print(f"   Pro escalation failed: {_pe}")

        return data

    async def generate():
        total = len(items)
        for i, item in enumerate(items):
            yield {"data": json.dumps({"type": "start", "lot": item.get("lot"), "index": i, "total": total})}
            try:
                images = []
                # Try to get image from uploaded PDF first
                if pdf_bytes and item.get("_page_start"):
                    images = await loop.run_in_executor(
                        executor, extract_page_image, pdf_bytes,
                        item["_page_start"], item.get("_page_end", item["_page_start"])
                    )
                # Fall back to fetching from stored PDF via scan_id
                if not images and item.get("_page_img"):
                    try:
                        img_url = item["_page_img"]
                        # Extract scan_id and img_index from URL like /api/auction/page-image/{scan_id}/{idx}
                        parts_url = img_url.strip("/").split("/")
                        if len(parts_url) >= 2:
                            sid = parts_url[-2]
                            idx = int(parts_url[-1])
                            stored_pdf = supabase.storage.from_("auction-pdfs").download(f"{sid}.pdf")
                            images = await loop.run_in_executor(
                                executor, lambda: extract_single_image(stored_pdf, idx)
                            )
                    except Exception as img_e:
                        print(f"Auto image fetch error: {img_e}")
                result = await loop.run_in_executor(executor, research_item, item, images)
                yield {"data": json.dumps({
                    "type": "result",
                    "lot": item.get("lot"),
                    "index": i,
                    "total": total,
                    "has_image": len(images) > 0,
                    **result
                })}
            except json.JSONDecodeError as e:
                print(f"JSON parse error for lot {item.get('lot')}: {e}")
                yield {"data": json.dumps({
                    "type": "result",
                    "lot": item.get("lot"),
                    "index": i,
                    "total": total,
                    "has_image": len(images) > 0,
                    "revised_value": item.get("your_value", 0),
                    "confidence": "low",
                    "comps": [],
                    "image_notes": "Research completed but response parsing failed",
                    "recommendation": "watch",
                    "rec_reason": "Could not parse research results — try again",
                    "notes": ""
                })}
            except Exception as e:
                print(f"Deep research error for lot {item.get('lot')}: {e}")
                yield {"data": json.dumps({
                    "type": "error",
                    "lot": item.get("lot"),
                    "index": i,
                    "total": total,
                    "error": str(e)
                })}
            await asyncio.sleep(0.1)
        yield {"data": json.dumps({"type": "done", "total": total})}

    return EventSourceResponse(generate())




@app.get("/api/auction/research-items/{scan_id}")
async def get_research_items(scan_id: str):
    """Return watchlisted items for a scan so any client can load them."""
    import json
    try:
        row = supabase.table("auction_research_sessions")             .select("items,results,title").eq("share_id", scan_id).single().execute()
        data = row.data
        return {
            "scan_id": scan_id,
            "title":   data.get("title",""),
            "items":   json.loads(data.get("items","[]")),
            "results": json.loads(data.get("results","{}")),
        }
    except Exception:
        return {"scan_id": scan_id, "items": [], "results": {}}


@app.post("/api/auction/save-research")
async def save_research(request: Request):
    import json, uuid
    body = await request.json()
    share_id = body.get("share_id") or str(uuid.uuid4())[:8]
    title    = body.get("title", "Auction Research")
    items    = body.get("items", [])
    results  = body.get("results", {})
    try:
        supabase.table("auction_research_sessions").upsert({
            "share_id": share_id,
            "title":    title,
            "items":    json.dumps(items),
            "results":  json.dumps(results),
        }, on_conflict="share_id").execute()
    except Exception as e:
        raise HTTPException(500, f"Save failed: {e}")
    return {"share_id": share_id}


@app.get("/api/auction/scans")
async def list_scans():
    try:
        res = supabase.table("auction_research_sessions")            .select("share_id, title, items, created_at")            .order("created_at", desc=True)            .limit(50)            .execute()
        scans = []
        for row in (res.data or []):
            import json as _j
            items = row.get("items") or []
            if isinstance(items, str):
                try: items = _j.loads(items)
                except: items = []
            scans.append({
                "id": row["share_id"],
                "name": row.get("title", row["share_id"]),
                "items": items,
                "ts": row.get("created_at", "")
            })
        return {"scans": scans}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/auction/scans/{scan_id}")
async def get_scan(scan_id: str):
    try:
        import json as _j
        res = supabase.table("auction_research_sessions")            .select("share_id, title, items")            .eq("share_id", scan_id)            .limit(1)            .execute()
        if not res.data:
            raise HTTPException(404, "Scan not found")
        row = res.data[0]
        items = row.get("items") or []
        if isinstance(items, str):
            try: items = _j.loads(items)
            except: items = []
        return {"id": row["share_id"], "name": row.get("title", row["share_id"]), "items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/api/auction/scans")
async def save_scan(request: Request):
    try:
        import json as _j
        body = await request.json()
        scan_id = body.get("id")
        name = body.get("name", scan_id)
        items = body.get("items", [])
        import json as _j
        supabase.table("auction_research_sessions").upsert({
            "share_id": scan_id,
            "title": name,
            "items": _j.dumps(items),
        }, on_conflict="share_id").execute()
        return {"ok": True, "id": scan_id}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.delete("/api/auction/scans/{scan_id}")
async def delete_scan(scan_id: str):
    try:
        supabase.table("auction_research_sessions")            .delete()            .eq("share_id", scan_id)            .execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/auction/load-research/{share_id}")
async def load_research(share_id: str):
    import json
    try:
        row = supabase.table("auction_research_sessions")             .select("*").eq("share_id", share_id).single().execute()
        data = row.data
        return {
            "share_id": data["share_id"],
            "title":    data.get("title", ""),
            "items":    json.loads(data.get("items", "[]")),
            "results":  json.loads(data.get("results", "{}")),
        }
    except Exception as e:
        raise HTTPException(404, f"Session not found: {e}")


@app.post("/api/auction/research-export")
async def research_export(request: Request):
    import io, json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    form = await request.form()
    items = json.loads(form.get("items", "[]"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Deep Research"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E2535")
    amber_fill = PatternFill("solid", fgColor="412402")
    green_fill = PatternFill("solid", fgColor="052E16")
    alt_fill = PatternFill("solid", fgColor="161B28")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="center")
    thin = Border(
        bottom=Side(style="thin", color="2D3348"),
        right=Side(style="thin", color="2D3348")
    )

    headers = ["Lot", "Title", "Original Value", "Revised Value", "Confidence", "Recommendation", "Notes", "Your Notes", "eBay Search"]
    col_widths = [8, 45, 15, 15, 12, 16, 40, 30, 20]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22

    for ri, item in enumerate(items, 2):
        row_data = [
            item.get("lot", ""),
            item.get("title", ""),
            item.get("original_value", 0),
            item.get("revised_value", 0),
            item.get("confidence", "").capitalize(),
            item.get("recommendation", "").capitalize(),
            item.get("rec_reason") or item.get("image_notes", ""),
            item.get("user_note", ""),
            "View eBay Sold"
        ]
        rec = item.get("recommendation", "").lower()
        fill = green_fill if rec == "buy" else amber_fill if rec == "watch" else (alt_fill if ri % 2 == 0 else None)

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill
            cell.border = thin
            cell.alignment = wrap if ci in (2, 7) else center
            if ci in (3, 4) and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0'
            if ci == 8 and item.get("ebay_search"):
                ws.cell(row=ri, column=ci).hyperlink = item["ebay_search"]
                ws.cell(row=ri, column=ci).font = Font(color="4A9EFF", underline="single")
        ws.row_dimensions[ri].height = 20

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=deep_research.xlsx"}
    )


@app.post("/api/auction/research-export")
async def research_export(request: Request):
    import io, json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    form = await request.form()
    items = json.loads(form.get("items", "[]"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Deep Research"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E2535")
    amber_fill = PatternFill("solid", fgColor="412402")
    green_fill = PatternFill("solid", fgColor="052E16")
    alt_fill = PatternFill("solid", fgColor="161B28")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="center")
    thin = Border(
        bottom=Side(style="thin", color="2D3348"),
        right=Side(style="thin", color="2D3348")
    )

    headers = ["Lot", "Title", "Original Value", "Revised Value", "Confidence", "Recommendation", "Notes", "eBay Search"]
    col_widths = [8, 45, 15, 15, 12, 16, 40, 20]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22

    for ri, item in enumerate(items, 2):
        row_data = [
            item.get("lot", ""),
            item.get("title", ""),
            item.get("original_value", 0),
            item.get("revised_value", 0),
            item.get("confidence", "").capitalize(),
            item.get("recommendation", "").capitalize(),
            item.get("rec_reason") or item.get("image_notes", ""),
            item.get("user_note", ""),
            "View eBay Sold"
        ]
        rec = item.get("recommendation", "").lower()
        fill = green_fill if rec == "buy" else amber_fill if rec == "watch" else (alt_fill if ri % 2 == 0 else None)

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill
            cell.border = thin
            cell.alignment = wrap if ci in (2, 7) else center
            if ci in (3, 4) and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0'
            if ci == 8 and item.get("ebay_search"):
                ws.cell(row=ri, column=ci).hyperlink = item["ebay_search"]
                ws.cell(row=ri, column=ci).font = Font(color="4A9EFF", underline="single")
        ws.row_dimensions[ri].height = 20

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=deep_research.xlsx"}
    )

# ── API: AUCTION DEEP RESEARCH ───────────────────────────────── #

class DeepResearch(BaseModel):
    title: str
    current_value: float = 0

@app.post("/api/auction/deep-research")
async def deep_research(body: DeepResearch):
    import os, asyncio, json
    from concurrent.futures import ThreadPoolExecutor
    import google.generativeai as genai
    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(400, "GEMINI_API_KEY not set")
    genai.configure(api_key=gemini_key)
    model = genai.GenerativeModel("gemini-2.5-flash")
    prompt = f"""You are an expert industrial equipment appraiser.
Research this auction item thoroughly: "{body.title}"
Current estimate: ${body.current_value}

Check eBay sold listings, industrial dealers, and recent auction results.
Assume working used condition.

Return ONLY a JSON object (no markdown):
{{"your_value": 5000, "notes": "Sold $4,500-$6,000 on eBay 2024"}}

your_value must be an integer."""

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)
    try:
        response = await loop.run_in_executor(executor, lambda: model.generate_content(prompt, generation_config={"max_output_tokens": 300}))
        raw = response.text.strip().replace("```json","").replace("```","").strip()
        try:
            data = json.loads(raw)
        except json.JSONDecodeError:
            from json_repair import repair_json
            data = json.loads(repair_json(raw))
        data["ai_overview_html"] = ai_overview_html
        data["grounding_sources"] = grounding_sources
        return data
    except Exception as e:
        raise HTTPException(500, str(e))


# ── API: EXCEL EXPORT ─────────────────────────────────────────── #

@app.post("/api/auction/export-excel")
async def export_excel(request: Request):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    body = await request.json()
    items = body.get("items", [])
    name = body.get("name", "Auction Scan")

    wb = Workbook()
    ws = wb.active
    ws.title = name[:31]

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1A1F2E")
    hv_fill = PatternFill("solid", fgColor="5C3A00")
    hv_font = Font(color="FAC775", bold=True)

    headers = ["Lot", "Title", "Est. Value", "Notes", "Deep Scan", "Watchlisted"]
    ws.append(headers)
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        val = int(item.get("your_value", 0) or 0)
        row = [
            str(item.get("lot", "")),
            str(item.get("title", "")),
            f"${val:,}",
            str(item.get("notes", "")),
            "Yes" if item.get("_deep") else "",
            "Yes" if item.get("_watch") else "",
        ]
        ws.append(row)
        if val >= 500:
            r = ws.max_row
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = hv_fill
                ws.cell(row=r, column=col).font = hv_font

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import datetime
    fn = f"auction_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"}
    )


# ── API: AUCTION PAGE IMAGE ───────────────────────────────────── #

@app.get("/api/auction/page-image/{scan_id}/{img_index}")
async def get_page_image(scan_id: str, img_index: int):
    import fitz
    from fastapi.responses import Response
    try:
        pdf_data = supabase.storage.from_("auction-pdfs").download(f"{scan_id}.pdf")
        doc = fitz.open(stream=pdf_data, filetype="pdf")
        # Collect all large embedded images (skip logos/watermarks < 5KB)
        all_images = []
        seen_xrefs = set()
        for page in doc:
            for img in page.get_images(full=True):
                xref = img[0]
                if xref in seen_xrefs:
                    continue
                seen_xrefs.add(xref)
                try:
                    base_image = doc.extract_image(xref)
                    if base_image and base_image.get("image") and len(base_image["image"]) > 8000:
                        all_images.append(base_image)
                except Exception as search_err:
                    print(f"   Search grounding failed: {search_err}")
                    pass
        # Fallback: render page and crop item image area for image-only PDFs
        if not all_images or img_index < 0 or img_index >= len(all_images):
            doc2 = fitz.open(stream=pdf_data, filetype="pdf")
            items_per_page = 3
            page_num = img_index // items_per_page
            slot = img_index % items_per_page
            if page_num >= len(doc2):
                page_num = len(doc2) - 1
            page = doc2[page_num]
            pw, ph = page.rect.width, page.rect.height
            slot_h = ph / items_per_page
            clip = fitz.Rect(0, slot * slot_h, pw * 0.28, (slot + 1) * slot_h)
            mat = fitz.Matrix(2.0, 2.0)
            pix = page.get_pixmap(matrix=mat, clip=clip)
            doc2.close()
            return Response(content=pix.tobytes("jpeg"), media_type="image/jpeg",
                          headers={"Cache-Control": "public, max-age=86400"})
        img_data = all_images[img_index]
        ext = img_data.get("ext", "jpeg")
        mime = "image/jpeg" if ext in ("jpg", "jpeg") else "image/" + ext
        return Response(content=img_data["image"], media_type=mime, headers={"Cache-Control": "public, max-age=86400"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))

# ── API: PDF AUCTION SCAN ─────────────────────────────────────── #

from sse_starlette.sse import EventSourceResponse

@app.post("/api/auction/scan-txt")
async def scan_txt_auction(file: UploadFile = File(...)):
    import os, json, asyncio
    import google.generativeai as genai

    contents = await file.read()
    text = contents.decode("utf-8", errors="ignore")

    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(400, "GEMINI_API_KEY not set")

    genai.configure(api_key=gemini_key)
    model = genai.GenerativeModel("gemini-2.5-flash")

    prompt_template = """You are a world-class auction appraiser with deep expertise in industrial equipment, lab instruments, and commercial goods.

Extract EVERY auction lot from this catalog text.

For each lot return a JSON object:
- lot: lot number as string
- title: full item title as written
- description: one sentence description
- estimate_low: integer dollar amount
- estimate_high: integer dollar amount
- your_value: integer (your single best estimate - total lot value)
- notes: brief market note with price source

PRICING RULES:
- All values MUST be plain integers (no $, no text)
- Base on ACTUAL used market values from eBay sold listings
- If no lots found, return: []
- Return ONLY a JSON array, no markdown"""

    # Split text into chunks of ~8000 chars
    chunk_size = 8000
    chunks = [text[i:i+chunk_size] for i in range(0, len(text), chunk_size)]
    if not chunks:
        chunks = [text]

    from sse_starlette.sse import EventSourceResponse

    async def generate():
        loop = asyncio.get_event_loop()
        from concurrent.futures import ThreadPoolExecutor
        executor = ThreadPoolExecutor(max_workers=1)
        all_items = []
        total = len(chunks)

        for i, chunk in enumerate(chunks):
            try:
                def call_gemini(c=chunk, idx=i):
                    response = model.generate_content(
                        [prompt_template, f"\nCATALOG SECTION {idx+1}/{total}:\n{c}"],
                        generation_config={"max_output_tokens": 16000}
                    )
                    return response.text

                raw = await loop.run_in_executor(executor, call_gemini)
                raw = " ".join(raw.splitlines())
                if "```" in raw:
                    raw = raw.split("```")[1]
                    if raw.startswith("json"):
                        raw = raw[4:]
                    raw = raw.strip()
                start = raw.find("[")
                end = raw.rfind("]") + 1
                if start >= 0 and end > start:
                    raw = raw[start:end]
                try:
                    items = json.loads(raw)
                except Exception:
                    from json_repair import repair_json
                    items = json.loads(repair_json(raw))
                all_items.extend(items)
                yield {
                    "data": json.dumps({
                        "chunk": i + 1,
                        "total_chunks": total,
                        "items": items,
                        "scan_id": None,
                        "done": False
                    }, separators=(',', ':'))
                }
            except Exception as e:
                print(f"TXT chunk {i+1} error: {e}")
            await asyncio.sleep(0.1)

        yield {"data": json.dumps({"done": True, "total": len(all_items), "scan_id": None})}

    return EventSourceResponse(generate())


@app.post("/api/auction/scan-pdf")
async def scan_pdf_auction(file: UploadFile = File(...)):
    import os, base64, json, fitz, asyncio, uuid
    import google.generativeai as genai

    contents = await file.read()
    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(400, "GEMINI_API_KEY not set")

    # Store PDF in Supabase for later image retrieval
    scan_id = str(uuid.uuid4())[:8]
    try:
        supabase.storage.from_("auction-pdfs").upload(
            path=f"{scan_id}.pdf",
            file=contents,
            file_options={"content-type": "application/pdf", "upsert": "true"}
        )
    except Exception as upload_err:
        print(f"PDF storage warning: {upload_err}")
        scan_id = None

    # Extract text chunks — fall back to image rendering for image-only PDFs
    try:
        doc = fitz.open(stream=contents, filetype="pdf")
        total_pages = len(doc)
        chunk_size = 2
        page_chunks = []
        page_images = []  # list of (page_num, jpeg_bytes) for image fallback
        for start in range(0, total_pages, chunk_size):
            end = min(start + chunk_size, total_pages)
            chunk_text = ""
            for page_num in range(start, end):
                chunk_text += doc[page_num].get_text() + "\n"
            if chunk_text.strip():
                page_chunks.append(chunk_text)

        # If no text found, render pages as images
        if not page_chunks:
            print(f"PDF has no text — switching to image scan ({total_pages} pages)")
            for page_num in range(total_pages):
                page = doc[page_num]
                mat = fitz.Matrix(2.0, 2.0)
                pix = page.get_pixmap(matrix=mat)
                page_images.append((page_num, pix.tobytes("jpeg")))
        doc.close()
    except Exception as e:
        raise HTTPException(500, f"PDF read error: {e}")

    genai.configure(api_key=gemini_key)
    model = genai.GenerativeModel("gemini-2.5-flash")

    prompt_template = """You are a world-class auction appraiser with deep expertise in industrial equipment, lab instruments, and commercial goods.

Extract EVERY auction lot from this catalog section.

For each lot return a JSON object:
- lot: lot number as string
- title: full item title as written
- description: one sentence description
- estimate_low: integer dollar amount
- estimate_high: integer dollar amount
- your_value: integer (your single best estimate - total lot value)
- notes: brief market note with price source

EXPERT AUCTION TITLE INTERPRETATION:
- Quantities: "(2)", "QTY (3)", "SET OF 4", "PAIR", "x3" = price TOTAL for ALL units combined
- Vague lots: "SHELF OF...", "PALLET OF...", "BOX OF..." = estimate total resale of all contents
- Condition notes like "AS-IS", "UNTESTED", "ACTIVATION NOT GUARANTEED" = still price as normal working condition
- Always search for the SPECIFIC brand + model for accurate pricing
- Ignore auction house names, catalog numbers, location references in titles

PRICING RULES:
- All values MUST be plain integers (no $, no text)
- Base on ACTUAL used market values from eBay sold listings
- If no lots found, return: []
- Return ONLY a JSON array, no markdown

Example: [{"lot":"5","title":"Oakton pH Meter","description":"Portable pH/ORP meter with case","estimate_low":80,"estimate_high":150,"your_value":100,"notes":"Sells $80-150 used on eBay"}]"""

    def call_gemini(chunk_text, i, total):
        response = model.generate_content(
            [prompt_template, f"\nCATALOG SECTION {i+1}/{total}:\n{chunk_text[:10000]}"],
            generation_config={"max_output_tokens": 16000}
        )
        return response.text

    def call_gemini_image(page_num, img_bytes, total):
        """Send a rendered page image to Gemini Vision for lot extraction."""
        print(f"   Image scan page {page_num+1}/{total}")
        img_prompt = prompt_template + f"\n\nThis is page {page_num+1} of {total} of an auction catalog. Extract all lots visible in this image."
        response = model.generate_content(
            [img_prompt, {"mime_type": "image/jpeg", "data": img_bytes}],
            generation_config={"max_output_tokens": 16000}
        )
        return response.text

    async def generate():
        import asyncio
        from concurrent.futures import ThreadPoolExecutor
        loop = asyncio.get_event_loop()
        executor = ThreadPoolExecutor(max_workers=1)
        all_items = []

        # Image-only PDF path
        if page_images and not page_chunks:
            total_chunks = len(page_images)
            for i, (page_num, img_bytes) in enumerate(page_images):
                try:
                    raw = await loop.run_in_executor(executor, call_gemini_image, page_num, img_bytes, total_chunks)
                    raw = " ".join(raw.splitlines())
                    if "```" in raw:
                        raw = raw.split("```")[1]
                        if raw.startswith("json"): raw = raw[4:]
                        raw = raw.strip()
                    start = raw.find("[")
                    end = raw.rfind("]") + 1
                    if start >= 0 and end > start:
                        raw = raw[start:end]
                    try:
                        items = json.loads(raw)
                    except Exception:
                        from json_repair import repair_json
                        items = json.loads(repair_json(raw))
                    base_idx = len(all_items)
                    all_items.extend(items)
                    for item in items:
                        item["_page_start"] = page_num + 1
                        item["_page_end"] = page_num + 1
                        if scan_id:
                            item["_page_img"] = f"/api/auction/page-image/{scan_id}/{base_idx + items.index(item)}"
                    yield {
                        "data": json.dumps({
                            "chunk": i + 1,
                            "total_chunks": total_chunks,
                            "items": items,
                            "scan_id": scan_id,
                            "done": False
                        }, separators=(',', ':'))
                    }
                except Exception as e:
                    print(f"Image page {page_num+1} error: {e}")
                await asyncio.sleep(0.1)
            yield {"data": json.dumps({"done": True, "total": len(all_items), "scan_id": scan_id})}
            return

        total_chunks = len(page_chunks)

        for i, chunk_text in enumerate(page_chunks):
            try:
                raw = await loop.run_in_executor(executor, call_gemini, chunk_text, i, total_chunks)
                raw = " ".join(raw.splitlines())
                if "```" in raw:
                    raw = raw.split("```")[1]
                    if raw.startswith("json"):
                        raw = raw[4:]
                    raw = raw.strip()
                start = raw.find("[")
                end = raw.rfind("]") + 1
                if start >= 0 and end > start:
                    raw = raw[start:end]
                try:
                    items = json.loads(raw)
                except Exception as search_err:
                    print(f"   Search grounding failed: {search_err}")
                    from json_repair import repair_json
                    items = json.loads(repair_json(raw))
                base_idx = len(all_items)
                all_items.extend(items)
                page_start = i * chunk_size + 1
                page_end = min((i + 1) * chunk_size, total_pages)
                for item in items:
                    item["_page_start"] = page_start
                    item["_page_end"] = page_end
                    if scan_id:
                        item["_page_img"] = f"/api/auction/page-image/{scan_id}/{base_idx + items.index(item)}"
                yield {
                    "data": json.dumps({
                        "chunk": i + 1,
                        "total_chunks": total_chunks,
                        "items": items,
                        "scan_id": scan_id,
                        "done": False
                    }, separators=(',', ':'))
                }
            except Exception as e:
                print(f"Chunk {i+1} error: {e}")
            await asyncio.sleep(0.1)

        yield {"data": json.dumps({"done": True, "total": len(all_items), "scan_id": scan_id})}

    return EventSourceResponse(generate())


def get_unmatched_photos():
    """Get all photos from storage that haven't been matched yet."""
    try:
        # Get all files in part-photos bucket
        res = supabase.storage.from_("part-photos").list()
        files = [f["name"] for f in (res or []) if f.get("name") and not f["name"].startswith(".")]
        return {"photos": files, "count": len(files)}
    except Exception as e:
        raise HTTPException(500, str(e))

class ScanPartsBody(BaseModel):
    part_numbers: list
    photo_ids:    list
    gemini_key:   Optional[str] = None

@app.post("/api/parts/scan")
async def scan_parts(body: ScanPartsBody):
    """
    Scan a batch of photos through Gemini Vision.
    For each photo, extract any visible part numbers and check against the list.
    Returns matches with confidence.
    """
    import threading
    gemini_key = body.gemini_key or os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(400, "Gemini API key required")
    if not body.part_numbers:
        raise HTTPException(400, "No part numbers provided")

    results = []
    part_set = [str(p).strip().upper() for p in body.part_numbers if str(p).strip()]

    try:
        import google.generativeai as genai
        genai.configure(api_key=gemini_key)
        model = genai.GenerativeModel("gemini-2.5-flash")
    except Exception:
        try:
            from google import genai as genai2
            from google.genai import types
            client = genai2.Client(api_key=gemini_key)
        except Exception as e:
            raise HTTPException(500, f"Gemini init failed: {e}")

    for photo_id in body.photo_ids[:50]:  # max 50 at a time
        try:
            # Download photo from Supabase
            img_bytes = supabase.storage.from_("part-photos").download(photo_id)
            if not img_bytes:
                continue

            # Build prompt
            parts_list = "\n".join(part_set[:200])
            prompt = f"""Examine this image carefully. 
Read ALL visible text including: part numbers, model numbers, serial numbers, labels, stamps, engravings, stickers, tags.

I am looking for matches to this list of part numbers:
{parts_list}

Return ONLY a JSON object:
{{
  "visible_text": ["list", "of", "all", "text", "you", "can", "read"],
  "part_numbers_found": ["any", "part", "numbers", "you", "see"],
  "matches": ["part numbers that exactly or closely match the search list"],
  "confidence": "high/medium/low",
  "notes": "brief note on what you see"
}}

If no text visible or no matches, still return the JSON with empty arrays."""

            try:
                import google.generativeai as genai
                genai.configure(api_key=gemini_key)
                model = genai.GenerativeModel("gemini-2.5-flash")
                import PIL.Image
                import io
                img = PIL.Image.open(io.BytesIO(img_bytes))
                response = model.generate_content([prompt, img])
                raw = response.text or ""
            except Exception as search_err:
                print(f"   Search grounding failed: {search_err}")
                try:
                    from google import genai as gc
                    from google.genai import types as gt
                    cl = gc.Client(api_key=gemini_key)
                    models = [m.name for m in cl.models.list()]
                    best = next((m for m in models if "gemini-2.5" in m or "gemini-2.0" in m), models[0] if models else "models/gemini-1.5-pro")
                    resp = cl.models.generate_content(
                        model=best,
                        contents=[gt.Part.from_bytes(data=img_bytes, mime_type="image/jpeg"), prompt]
                    )
                    raw = resp.text or ""
                except Exception as e2:
                    results.append({"photo_id": photo_id, "error": str(e2), "matches": []})
                    continue

            import re, json
            raw = re.sub(r"^```[a-z]*\n?", "", raw, flags=re.IGNORECASE)
            raw = re.sub(r"\n?```$", "", raw).strip()
            jm = re.search(r'\{.*\}', raw, re.DOTALL)
            if jm:
                data = json.loads(jm.group())
                results.append({
                    "photo_id":        photo_id,
                    "url":             photo_url(photo_id, thumb=True),
                    "full_url":        photo_url(photo_id),
                    "visible_text":    data.get("visible_text", []),
                    "part_numbers_found": data.get("part_numbers_found", []),
                    "matches":         data.get("matches", []),
                    "confidence":      data.get("confidence", ""),
                    "notes":           data.get("notes", ""),
                    "has_match":       len(data.get("matches", [])) > 0,
                })
            else:
                results.append({"photo_id": photo_id, "matches": [], "notes": "Could not parse response"})

        except Exception as e:
            results.append({"photo_id": photo_id, "error": str(e), "matches": []})

    matches    = [r for r in results if r.get("has_match")]
    no_matches = [r for r in results if not r.get("has_match")]
    return {
        "results":     results,
        "matches":     matches,
        "no_matches":  no_matches,
        "match_count": len(matches),
        "scanned":     len(results),
    }

# ── SETTINGS ──────────────────────────────────────────────────── #

@app.get("/api/settings")
async def get_settings():
    try:
        res = supabase.table("app_settings").select("*").execute()
        return {row["key"]: row["value"] for row in (res.data or [])}
    except Exception:
        return {}

class SaveSetting(BaseModel):
    key:   str
    value: str

@app.post("/api/settings")
async def save_setting(body: SaveSetting):
    try:
        supabase.table("app_settings").upsert({"key": body.key, "value": body.value}).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))
import hashlib, secrets

def hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    hashed = hashlib.sha256((salt + password).encode()).hexdigest()
    return f"{salt}:{hashed}"

def verify_password(password: str, stored: str) -> bool:
    try:
        salt, hashed = stored.split(":")
        return hashlib.sha256((salt + password).encode()).hexdigest() == hashed
    except Exception:
        return False

def get_business_id(request: Request):
    """Get business_id from session cookie."""
    session = request.cookies.get("session_id")
    if not session:
        return None
    try:
        res = supabase.table("sessions").select("business_id").eq("token", session).execute()
        if res.data:
            return res.data[0]["business_id"]
    except Exception:
        pass
    return None

@app.get("/login")
async def login_page(request: Request, error: str = ""):
    return templates.TemplateResponse("login.html", {"request": request, "error": error})

@app.post("/login")
async def login_submit(request: Request):
    form = await request.form()
    email = str(form.get("email", "")).strip().lower()
    password = str(form.get("password", ""))
    try:
        res = supabase.table("businesses").select("id,password_hash").eq("email", email).execute()
        if not res.data:
            return templates.TemplateResponse("login.html", {"request": request, "error": "Invalid email or password"})
        biz = res.data[0]
        if not verify_password(password, biz["password_hash"]):
            return templates.TemplateResponse("login.html", {"request": request, "error": "Invalid email or password"})
        token = secrets.token_hex(32)
        supabase.table("sessions").insert({"token": token, "business_id": biz["id"]}).execute()
        from fastapi.responses import RedirectResponse
        resp = RedirectResponse("/", status_code=302)
        resp.set_cookie("session_id", token, httponly=True, max_age=60*60*24*30)
        return resp
    except Exception as e:
        return templates.TemplateResponse("login.html", {"request": request, "error": f"Login failed: {e}"})

@app.get("/register")
async def register_page(request: Request, error: str = ""):
    return templates.TemplateResponse("register.html", {"request": request, "error": error})

@app.post("/register")
async def register_submit(request: Request):
    form = await request.form()
    business_name = str(form.get("business_name", "")).strip()
    email = str(form.get("email", "")).strip().lower()
    password = str(form.get("password", ""))
    if not business_name or not email or not password:
        return templates.TemplateResponse("register.html", {"request": request, "error": "All fields required"})
    if len(password) < 8:
        return templates.TemplateResponse("register.html", {"request": request, "error": "Password must be at least 8 characters"})
    try:
        existing = supabase.table("businesses").select("id").eq("email", email).execute()
        if existing.data:
            return templates.TemplateResponse("register.html", {"request": request, "error": "Email already registered"})
        password_hash = hash_password(password)
        res = supabase.table("businesses").insert({
            "name": business_name,
            "email": email,
            "password_hash": password_hash
        }).execute()
        business_id = res.data[0]["id"]
        token = secrets.token_hex(32)
        supabase.table("sessions").insert({"token": token, "business_id": business_id}).execute()
        from fastapi.responses import RedirectResponse
        resp = RedirectResponse("/welcome", status_code=302)
        resp.set_cookie("session_id", token, httponly=True, max_age=60*60*24*30)
        return resp
    except Exception as e:
        return templates.TemplateResponse("register.html", {"request": request, "error": f"Registration failed: {e}"})

@app.get("/logout")
async def logout(request: Request):
    from fastapi.responses import RedirectResponse
    token = request.cookies.get("session_id")
    if token:
        try:
            supabase.table("sessions").delete().eq("token", token).execute()
        except Exception:
            pass
    resp = RedirectResponse("/login", status_code=302)
    resp.delete_cookie("session_id")
    return resp





@app.get("/api/admin/bad-reports")
async def list_bad_reports(status: str = None):
    """Admin: list bad scan reports."""
    try:
        q = supabase.table("bad_scan_reports").select("*").order("reported_at", desc=True)
        if status:
            q = q.eq("status", status)
        result = q.execute()

        for r in result.data:
            r["photo_urls"] = [
                f"{SUPABASE_URL}/storage/v1/object/public/part-photos/{pid}"
                for pid in (r.get("photo_ids") or [])
            ]

        counts = {"new": 0, "reviewed": 0, "resolved": 0}
        all_rows = supabase.table("bad_scan_reports").select("status").execute()
        for r in all_rows.data:
            s = r.get("status", "new")
            if s in counts:
                counts[s] += 1

        return {"reports": result.data, "counts": counts}
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))


@app.patch("/api/admin/bad-reports/{report_id}")
async def update_bad_report(report_id: str, request: Request):
    """Admin: update report status."""
    try:
        data = await request.json()
        if data.get("status") not in ("new", "reviewed", "resolved"):
            raise HTTPException(400, "Invalid status")
        supabase.table("bad_scan_reports").update({"status": data["status"]}).eq("id", report_id).execute()
        return {"ok": True}
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))




@app.post("/api/admin/bad-reports/{report_id}/send-reply")
async def send_bad_report_reply(report_id: str, request: Request):
    """Admin sends an email reply to the reporter via Resend, then marks resolved."""
    try:
        body = await request.json()
        to_email = body.get("to", "").strip()
        subject = body.get("subject", "").strip()
        message = body.get("body", "").strip()
        if not to_email or not subject or not message:
            raise HTTPException(400, "to, subject, and body are required")

        import resend
        resend.api_key = os.getenv("RESEND_API_KEY", "")
        # Convert plain text to HTML (preserve line breaks)
        html_body = message.replace("\n", "<br>")
        email_html = f'''<div style="font-family:-apple-system,sans-serif;max-width:560px;padding:16px;color:#1f2937;line-height:1.6;">{html_body}</div>'''
        try:
            resend.Emails.send({
                "from": "Lister AI Reports <reports@reselljunkie.com>",
                "to": to_email,
                "subject": subject,
                "html": email_html
            })
        except Exception:
            resend.Emails.send({
                "from": "Lister AI <onboarding@resend.dev>",
                "to": to_email,
                "subject": subject,
                "html": email_html
            })

        # Mark resolved
        supabase.table("bad_scan_reports").update({"status": "resolved"}).eq("id", report_id).execute()
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))

@app.delete("/api/admin/bad-reports/{report_id}")
async def delete_bad_report(report_id: str):
    """Admin: delete a report."""
    try:
        supabase.table("bad_scan_reports").delete().eq("id", report_id).execute()
        return {"ok": True}
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))


# ─── Account Management ───────────────────────────────────────

@app.get("/api/account/info")
async def get_account_info(request: Request):
    """Return current user's account: name, email, plan, scan usage."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        biz = supabase.table("businesses").select("*").eq("id", business_id).execute()
        if not biz.data:
            raise HTTPException(404, "Not found")
        b = biz.data[0]
        # Infer plan from scan_limit
        limit = b.get("scan_limit", 25)
        plan_map = {25: "Free Trial", 100: "Starter", 500: "Growth", 1000: "Pro"}
        plan = plan_map.get(limit, "Custom")
        return {
            "name": b.get("name", ""),
            "email": b.get("email", ""),
            "plan": plan,
            "scan_limit": limit,
            "scan_count": b.get("scan_count", 0),
            "created_at": b.get("created_at", ""),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.patch("/api/account/info")
async def update_account_info(request: Request):
    """Update business name and/or email."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        updates = {}
        if "name" in body:
            updates["name"] = str(body["name"]).strip()
        if "email" in body:
            new_email = str(body["email"]).strip().lower()
            # Check no other business has this email
            existing = supabase.table("businesses").select("id").eq("email", new_email).execute()
            if existing.data and existing.data[0]["id"] != business_id:
                raise HTTPException(400, "Email already in use")
            updates["email"] = new_email
        if not updates:
            raise HTTPException(400, "No fields to update")
        supabase.table("businesses").update(updates).eq("id", business_id).execute()
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/account/change-plan")
async def change_account_plan(request: Request):
    """Change plan tier (no payment processing — internal/dev only)."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        plan = body.get("plan", "")
        plan_limits = {"Free Trial": 25, "Starter": 100, "Growth": 500, "Pro": 1000}
        if plan not in plan_limits:
            raise HTTPException(400, "Invalid plan")
        supabase.table("businesses").update({"scan_limit": plan_limits[plan]}).eq("id", business_id).execute()
        return {"ok": True, "plan": plan, "scan_limit": plan_limits[plan]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/account/contact-support")
async def contact_support(request: Request):
    """User submits a support message — sent via Resend to admin."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        message = str(body.get("message", "")).strip()
        topic = str(body.get("topic", "General")).strip()
        if not message:
            raise HTTPException(400, "Message required")
        biz = supabase.table("businesses").select("name,email").eq("id", business_id).execute()
        biz_data = biz.data[0] if biz.data else {}

        import resend
        resend.api_key = os.getenv("RESEND_API_KEY", "")
        html = f"""
<div style="font-family:-apple-system,sans-serif;max-width:560px;padding:24px;background:#0a0c10;color:#f0f2f5;">
  <h1 style="color:#3b82f6;font-size:20px;margin-bottom:16px;">📩 Support Request</h1>
  <p style="margin:6px 0;"><strong>From:</strong> {biz_data.get('name','Unknown')} ({biz_data.get('email','')})</p>
  <p style="margin:6px 0;"><strong>Topic:</strong> {topic}</p>
  <p style="margin:6px 0;"><strong>Business ID:</strong> {business_id}</p>
  <div style="background:#111418;border:1px solid #1a1f2e;border-radius:10px;padding:18px;margin-top:16px;">
    <div style="white-space:pre-wrap;font-size:14px;line-height:1.6;">{message}</div>
  </div>
</div>
"""
        try:
            resend.Emails.send({
                "from": "Lister AI Support <support@reselljunkie.com>",
                "reply_to": biz_data.get("email", ""),
                "to": "sebastian@lilbeecreative.com",
                "subject": f"Support: {topic} — {biz_data.get('name','Unknown')}",
                "html": html
            })
        except Exception:
            resend.Emails.send({
                "from": "Lister AI <onboarding@resend.dev>",
                "to": "sebastian@lilbeecreative.com",
                "subject": f"Support: {topic}",
                "html": html
            })
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))


# ─── Support Messaging ─────────────────────────────────────────

def _get_or_create_thread(business_id):
    """Return existing open thread for business, or create a new one."""
    res = supabase.table("support_threads").select("*").eq("business_id", business_id).order("last_message_at", desc=True).limit(1).execute()
    if res.data:
        return res.data[0]
    new = supabase.table("support_threads").insert({"business_id": business_id}).execute()
    return new.data[0]


# ─── User-side messaging ───

@app.get("/api/support/thread")
async def user_get_thread(request: Request):
    """User: get their own thread + messages."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        thread = _get_or_create_thread(business_id)
        msgs = supabase.table("support_messages").select("*").eq("thread_id", thread["id"]).order("created_at").execute()
        # Mark as read for user
        supabase.table("support_threads").update({"unread_for_user": False}).eq("id", thread["id"]).execute()
        return {"thread": thread, "messages": msgs.data or []}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/support/send")
async def user_send_message(request: Request):
    """User: send a message to support."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        text = str(body.get("body", "")).strip()
        if not text:
            raise HTTPException(400, "Message required")
        thread = _get_or_create_thread(business_id)
        from datetime import datetime, timezone
        now = datetime.now(timezone.utc).isoformat()
        supabase.table("support_messages").insert({
            "thread_id": thread["id"],
            "sender": "user",
            "body": text
        }).execute()
        supabase.table("support_threads").update({
            "last_message_at": now,
            "unread_for_admin": True
        }).eq("id", thread["id"]).execute()
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/support/unread")
async def user_unread_check(request: Request):
    """User: check if they have unread admin messages."""
    business_id = require_auth(request)
    if not business_id:
        return {"unread": False}
    try:
        res = supabase.table("support_threads").select("id,unread_for_user").eq("business_id", business_id).eq("unread_for_user", True).limit(1).execute()
        return {"unread": bool(res.data)}
    except Exception:
        return {"unread": False}


# ─── Admin-side messaging ───

@app.get("/api/admin/threads")
async def admin_list_threads(request: Request):
    """Admin: list all support threads with last message preview + business info."""
    try:
        threads = supabase.table("support_threads").select("*").order("last_message_at", desc=True).execute()
        out = []
        unread_count = 0
        for t in (threads.data or []):
            biz = supabase.table("businesses").select("name,email").eq("id", t["business_id"]).limit(1).execute()
            biz_data = biz.data[0] if biz.data else {}
            last_msg = supabase.table("support_messages").select("body,sender,created_at").eq("thread_id", t["id"]).order("created_at", desc=True).limit(1).execute()
            preview = last_msg.data[0] if last_msg.data else None
            if t.get("unread_for_admin"):
                unread_count += 1
            out.append({
                **t,
                "business_name": biz_data.get("name", "Unknown"),
                "business_email": biz_data.get("email", ""),
                "last_message": preview,
            })
        return {"threads": out, "unread_count": unread_count}
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))


@app.get("/api/admin/threads/{thread_id}")
async def admin_get_thread(thread_id: str):
    """Admin: get full thread + mark as read for admin."""
    try:
        thread = supabase.table("support_threads").select("*").eq("id", thread_id).limit(1).execute()
        if not thread.data:
            raise HTTPException(404, "Not found")
        msgs = supabase.table("support_messages").select("*").eq("thread_id", thread_id).order("created_at").execute()
        biz = supabase.table("businesses").select("name,email").eq("id", thread.data[0]["business_id"]).limit(1).execute()
        supabase.table("support_threads").update({"unread_for_admin": False}).eq("id", thread_id).execute()
        return {
            "thread": thread.data[0],
            "messages": msgs.data or [],
            "business": biz.data[0] if biz.data else {}
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/admin/threads/{thread_id}/reply")
async def admin_reply(thread_id: str, request: Request):
    """Admin: reply to a thread."""
    try:
        body = await request.json()
        text = str(body.get("body", "")).strip()
        if not text:
            raise HTTPException(400, "Message required")
        from datetime import datetime, timezone
        now = datetime.now(timezone.utc).isoformat()
        supabase.table("support_messages").insert({
            "thread_id": thread_id,
            "sender": "admin",
            "body": text
        }).execute()
        supabase.table("support_threads").update({
            "last_message_at": now,
            "unread_for_user": True
        }).eq("id", thread_id).execute()
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/admin/businesses/{business_id}/start-thread")
async def admin_start_thread(business_id: str, request: Request):
    """Admin: start (or get) a thread with a business and send first message."""
    try:
        body = await request.json()
        text = str(body.get("body", "")).strip()
        if not text:
            raise HTTPException(400, "Message required")
        thread = _get_or_create_thread(business_id)
        from datetime import datetime, timezone
        now = datetime.now(timezone.utc).isoformat()
        supabase.table("support_messages").insert({
            "thread_id": thread["id"],
            "sender": "admin",
            "body": text
        }).execute()
        supabase.table("support_threads").update({
            "last_message_at": now,
            "unread_for_user": True
        }).eq("id", thread["id"]).execute()
        return {"ok": True, "thread_id": thread["id"]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# ─── Cost Tracking ─────────────────────────────────────────────

@app.patch("/api/listings/{listing_id}/cost")
async def update_listing_cost(listing_id: int, request: Request):
    """Update the cost of a single listing."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        cost = float(body.get("cost", 0) or 0)
        if cost < 0:
            raise HTTPException(400, "Cost cannot be negative")
        supabase.table("listings").update({"cost": cost}).eq("id", listing_id).eq("business_id", business_id).execute()
        return {"ok": True, "cost": cost}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/listings/lot-cost")
async def apply_lot_cost(request: Request):
    """Distribute a total lot cost across selected listings."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        listing_ids = body.get("listing_ids", [])
        total_cost = float(body.get("total_cost", 0) or 0)
        method = body.get("method", "even")  # 'even' or 'weighted'
        if not listing_ids or total_cost <= 0:
            raise HTTPException(400, "listing_ids and total_cost required")

        if method == "weighted":
            # Distribute proportionally to listed price
            res = supabase.table("listings").select("id,price").in_("id", listing_ids).eq("business_id", business_id).execute()
            items = res.data or []
            total_listed = sum(float(it.get("price") or 0) for it in items)
            if total_listed <= 0:
                # Fall back to even split
                method = "even"
            else:
                for it in items:
                    portion = (float(it.get("price") or 0) / total_listed) * total_cost
                    supabase.table("listings").update({"cost": round(portion, 2)}).eq("id", it["id"]).eq("business_id", business_id).execute()
                return {"ok": True, "method": "weighted", "applied_to": len(items)}

        # Even split
        per_item = round(total_cost / len(listing_ids), 2)
        for lid in listing_ids:
            supabase.table("listings").update({"cost": per_item}).eq("id", lid).eq("business_id", business_id).execute()
        return {"ok": True, "method": "even", "per_item": per_item, "applied_to": len(listing_ids)}
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))



@app.get("/api/home/overview")
async def home_overview(request: Request):
    """Single endpoint that returns all metrics for the new Home dashboard."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        from datetime import datetime, timezone, timedelta
        now = datetime.now(timezone.utc)
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        # Inventory value + count — read from inventory table (same source as Inventory page)
        inv = (supabase.table("inventory")
            .select("price, status")
            .eq("business_id", business_id)
            .execute())
        inv_items = inv.data or []
        inventory_value = 0.0
        total_items = 0
        for it in inv_items:
            if it.get("status") != "sold":
                inventory_value += float(it.get("price") or 0)
                total_items += 1

        # Sold this month — from inventory table (sold_date set this month)
        try:
            sold_res = (supabase.table("inventory")
                .select("sold_price, sold_date")
                .eq("business_id", business_id)
                .gte("sold_date", month_start.date().isoformat())
                .execute())
            sold_items = sold_res.data or []
            sold_mtd = sum(float(s.get("sold_price") or 0) for s in sold_items)
            sold_count = len(sold_items)
        except Exception:
            sold_mtd = 0.0
            sold_count = 0

        # Spent this month — from expenses table
        try:
            exp_res = (supabase.table("expenses")
                .select("amount")
                .eq("business_id", business_id)
                .gte("expense_date", month_start.date().isoformat())
                .execute())
            spent_mtd = sum(float(e.get("amount") or 0) for e in (exp_res.data or []))
        except Exception:
            spent_mtd = 0.0

        # Scan limits + business name
        biz = supabase.table("businesses").select(
            "name, scan_count, scan_limit"
        ).eq("id", business_id).limit(1).execute()
        b = biz.data[0] if biz.data else {}

        # Connection statuses from token tables
        try:
            ebay_tok = supabase.table("ebay_tokens").select("id").eq("business_id", business_id).limit(1).execute()
            ebay_connected = bool(ebay_tok.data)
        except Exception:
            ebay_connected = False
        try:
            shop_tok = supabase.table("shopify_tokens").select("id").eq("business_id", business_id).limit(1).execute()
            shopify_connected = bool(shop_tok.data)
        except Exception:
            shopify_connected = False

        return {
            "business_name": b.get("name", ""),
            "inventory_value": round(inventory_value, 2),
            "total_items": total_items,
            "sold_mtd": round(sold_mtd, 2),
            "sold_count_mtd": sold_count,
            "spent_mtd": round(spent_mtd, 2),
            "scan_count": b.get("scan_count", 0),
            "scan_limit": b.get("scan_limit", 25),
            "ebay_connected": ebay_connected,
            "shopify_connected": shopify_connected,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/dashboard/summary")
async def dashboard_summary(request: Request, period: str = "month"):
    """Return profit/cost/listed totals for a time period."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        from datetime import datetime, timezone, timedelta
        now = datetime.now(timezone.utc)
        if period == "day":
            since = now - timedelta(days=1)
        elif period == "week":
            since = now - timedelta(days=7)
        elif period == "month":
            since = now - timedelta(days=30)
        elif period == "3month":
            since = now - timedelta(days=90)
        elif period == "6month":
            since = now - timedelta(days=180)
        elif period == "year":
            since = now - timedelta(days=365)
        elif period == "lifetime":
            since = now - timedelta(days=365*20)
        else:
            since = now - timedelta(days=30)

        # Get all active listings in window
        res = (supabase.table("listings")
            .select("id,price,cost,quantity,status")
            .eq("business_id", business_id)
            .neq("status", "archived")
            .gte("created_at", since.isoformat())
            .execute())
        items = res.data or []

        total_listed = sum(float(it.get("price") or 0) for it in items)
        total_cost = sum(float(it.get("cost") or 0) for it in items)
        total_units = sum(int(it.get("quantity") or 1) for it in items)
        total_items = len(items)
        profit = total_listed - total_cost

        # Also return cost_tracking_enabled status
        biz = supabase.table("businesses").select("cost_tracking_enabled,scan_count,scan_limit").eq("id", business_id).limit(1).execute()
        biz_data = biz.data[0] if biz.data else {}

        return {
            "period": period,
            "items": total_items,
            "units": total_units,
            "total_cost": round(total_cost, 2),
            "total_listed": round(total_listed, 2),
            "profit": round(profit, 2),
            "cost_tracking_enabled": biz_data.get("cost_tracking_enabled", True),
            "scan_count": biz_data.get("scan_count", 0),
            "scan_limit": biz_data.get("scan_limit", 25),
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))


@app.patch("/api/account/cost-tracking")
async def toggle_cost_tracking(request: Request):
    """Turn cost tracking on or off for the current business."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        enabled = bool(body.get("enabled", True))
        supabase.table("businesses").update({"cost_tracking_enabled": enabled}).eq("id", business_id).execute()
        return {"ok": True, "enabled": enabled}
    except Exception as e:
        raise HTTPException(500, str(e))


# ─── Inventory ──────────────────────────────────────────────────

GENERAL_CATEGORY_MAP = {
    # Maps eBay category fragments → broad buckets
    "electronics": "Electronics", "phone": "Electronics", "computer": "Electronics",
    "camera": "Electronics", "audio": "Electronics", "tv": "Electronics",
    "clothing": "Clothing", "shoes": "Clothing", "apparel": "Clothing", "jewelry": "Clothing",
    "tops": "Clothing", "shirt": "Clothing", "pant": "Clothing", "dress": "Clothing",
    "tool": "Tools", "hardware": "Tools", "industrial": "Tools",
    "home": "Home", "kitchen": "Home", "garden": "Home", "furniture": "Home",
    "toy": "Toys", "game": "Toys", "doll": "Toys",
    "book": "Books", "magazine": "Books",
    "auto": "Auto", "car": "Auto", "motor": "Auto", "vehicle": "Auto",
    "sport": "Sports", "fitness": "Sports", "bike": "Sports", "outdoor": "Sports",
    "collectible": "Collectibles", "antique": "Collectibles", "art": "Collectibles", "coin": "Collectibles",
    "health": "Health", "beauty": "Health",
}


def derive_general_category(ebay_category: str) -> str:
    if not ebay_category:
        return "Other"
    lower = ebay_category.lower()
    for key, bucket in GENERAL_CATEGORY_MAP.items():
        if key in lower:
            return bucket
    return "Other"


def add_to_inventory(business_id: str, listing: dict):
    """Insert a listing into the inventory table. Called automatically after scan."""
    try:
        # Skip if already in inventory (idempotent)
        existing = supabase.table("inventory").select("id").eq("source_listing_id", listing["id"]).eq("business_id", business_id).limit(1).execute()
        if existing.data:
            return existing.data[0]["id"]
        # Get all photos
        photo_ids = []
        if listing.get("photo_id"):
            try:
                gp = supabase.table("group_photos").select("group_id").eq("photo_id", listing["photo_id"]).limit(1).execute()
                if gp.data:
                    gid = gp.data[0]["group_id"]
                    all_gp = supabase.table("group_photos").select("photo_id").eq("group_id", gid).execute()
                    photo_ids = [r["photo_id"] for r in (all_gp.data or [])]
            except Exception:
                pass
            if not photo_ids:
                photo_ids = [listing["photo_id"]]
        new_inv = {
            "business_id": business_id,
            "source_listing_id": listing.get("id"),
            "title": listing.get("title", ""),
            "photo_id": listing.get("photo_id"),
            "all_photo_ids": photo_ids,
            "price": listing.get("price") or 0,
            "price_used": listing.get("price_used") or 0,
            "price_new": listing.get("price_new") or 0,
            "cost": listing.get("cost") or 0,
            "condition": listing.get("condition", "used"),
            "quantity": listing.get("quantity") or 1,
            "ebay_category": listing.get("ebay_category"),
            "general_category": derive_general_category(listing.get("ebay_category", "")),
        }
        result = supabase.table("inventory").insert(new_inv).execute()
        return result.data[0]["id"] if result.data else None
    except Exception as e:
        print(f"add_to_inventory failed: {e}")
        return None


@app.get("/api/inventory")
async def get_inventory(request: Request, search: str = "", status: str = "", location: str = "", category: str = ""):
    """List inventory items with optional filters. Auto-syncs scanned listings on each load."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        # Auto-backfill: any scanned listings not yet in inventory get added
        try:
            scanned = supabase.table("listings").select("*").eq("business_id", business_id).eq("status", "scanned").execute()
            existing_ids = supabase.table("inventory").select("source_listing_id").eq("business_id", business_id).execute()
            existing_set = {r["source_listing_id"] for r in (existing_ids.data or []) if r.get("source_listing_id")}
            for l in (scanned.data or []):
                if l["id"] not in existing_set:
                    add_to_inventory(business_id, l)
        except Exception as _e:
            print(f"inventory auto-sync warning: {_e}")
        q = supabase.table("inventory").select("*").eq("business_id", business_id).order("scanned_at", desc=True)
        if status:
            q = q.eq("status", status)
        if location:
            q = q.eq("storage_location", location)
        if category:
            q = q.eq("general_category", category)
        result = q.execute()
        items = result.data or []

        # Client-side search across title and barcode_id
        if search:
            s = search.lower()
            items = [i for i in items if s in (i.get("title") or "").lower() or s in (i.get("barcode_id") or "").lower()]

        # Build photo URLs
        for it in items:
            it["photo_url"] = f"{SUPABASE_URL}/storage/v1/object/public/part-photos/{it['photo_id']}" if it.get("photo_id") else ""
            it["all_photo_urls"] = [f"{SUPABASE_URL}/storage/v1/object/public/part-photos/{p}" for p in (it.get("all_photo_ids") or [])]

        # Aggregate stats
        total_value = sum(float(i.get("price") or 0) for i in items)
        total_cost = sum(float(i.get("cost") or 0) for i in items)
        in_stock = sum(1 for i in items if i.get("status") == "in_stock")
        listed = sum(1 for i in items if i.get("status") == "listed")
        sold = sum(1 for i in items if i.get("status") == "sold")

        return {
            "items": items,
            "stats": {
                "total": len(items),
                "in_stock": in_stock,
                "listed": listed,
                "sold": sold,
                "total_value": round(total_value, 2),
                "total_cost": round(total_cost, 2),
            }
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))


@app.patch("/api/inventory/{inv_id}")
async def update_inventory(inv_id: str, request: Request):
    """Update fields on an inventory item."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        allowed = ("title", "price", "cost", "condition", "quantity", "general_category",
                   "storage_location", "barcode_id", "status", "sold_price", "sold_date", "sold_location")
        updates = {k: v for k, v in body.items() if k in allowed}
        if not updates:
            raise HTTPException(400, "No valid fields")
        from datetime import datetime, timezone
        updates["updated_at"] = datetime.now(timezone.utc).isoformat()
        supabase.table("inventory").update(updates).eq("id", inv_id).eq("business_id", business_id).execute()
        # Track storage location for autocomplete
        if "storage_location" in updates and updates["storage_location"]:
            try:
                existing = supabase.table("storage_locations").select("id,use_count").eq("business_id", business_id).eq("name", updates["storage_location"]).limit(1).execute()
                if existing.data:
                    supabase.table("storage_locations").update({"use_count": (existing.data[0].get("use_count") or 0) + 1}).eq("id", existing.data[0]["id"]).execute()
                else:
                    supabase.table("storage_locations").insert({"business_id": business_id, "name": updates["storage_location"]}).execute()
            except Exception:
                pass
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/api/inventory/{inv_id}")
async def delete_inventory(inv_id: str, request: Request):
    """Permanently delete an inventory item."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        supabase.table("inventory").delete().eq("id", inv_id).eq("business_id", business_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/inventory/bulk-update")
async def bulk_update_inventory(request: Request):
    """Apply same field update to many inventory items."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        ids = body.get("ids", [])
        updates = {k: v for k, v in body.items() if k in ("storage_location", "status", "general_category")}
        if not ids or not updates:
            raise HTTPException(400, "ids and updates required")
        for inv_id in ids:
            supabase.table("inventory").update(updates).eq("id", inv_id).eq("business_id", business_id).execute()
        # Track location for autocomplete
        if updates.get("storage_location"):
            try:
                name = updates["storage_location"]
                existing = supabase.table("storage_locations").select("id,use_count").eq("business_id", business_id).eq("name", name).limit(1).execute()
                if existing.data:
                    supabase.table("storage_locations").update({"use_count": (existing.data[0].get("use_count") or 0) + len(ids)}).eq("id", existing.data[0]["id"]).execute()
                else:
                    supabase.table("storage_locations").insert({"business_id": business_id, "name": name, "use_count": len(ids)}).execute()
            except Exception:
                pass
        return {"ok": True, "updated": len(ids)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/inventory/locations")
async def list_locations(request: Request):
    """Return user's storage locations sorted by usage."""
    business_id = require_auth(request)
    if not business_id:
        return {"locations": []}
    try:
        result = supabase.table("storage_locations").select("name,use_count").eq("business_id", business_id).order("use_count", desc=True).execute()
        return {"locations": [r["name"] for r in (result.data or [])]}
    except Exception:
        return {"locations": []}


@app.post("/api/inventory/sync-from-listings")
async def sync_inventory_from_listings(request: Request):
    """Backfill inventory from existing listings (one-time op for current customers)."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        listings = supabase.table("listings").select("*").eq("business_id", business_id).neq("status", "archived").execute()
        added = 0
        for l in (listings.data or []):
            if add_to_inventory(business_id, l):
                added += 1
        return {"ok": True, "added": added}
    except Exception as e:
        raise HTTPException(500, str(e))


# ─── Multi-Item Photo Detection ───────────────────────────────

MULTI_ITEM_MIN_TIER_LIMIT = 500  # Growth plan or higher
MULTI_ITEM_CAP = 20


@app.post("/api/multi-scan/detect")
async def multi_scan_detect(request: Request):
    """Detect items in a single photo using Gemini, return bounding boxes for review."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        # Gate behind plan tier (Growth = scan_limit 500)
        biz = supabase.table("businesses").select("scan_limit,scan_count,is_admin").eq("id", business_id).limit(1).execute()
        if not biz.data:
            raise HTTPException(404, "Business not found")
        b = biz.data[0]
        if not b.get("is_admin") and (b.get("scan_limit") or 0) < MULTI_ITEM_MIN_TIER_LIMIT:
            raise HTTPException(402, "Multi-item scanning requires Growth plan or higher")

        form = await request.form()
        file = form.get("file")
        if not file:
            raise HTTPException(400, "No file")
        contents = await file.read()

        # Send full photo to Gemini for detection
        import google.generativeai as genai
        from PIL import Image
        import io as _io
        gemini_key = os.getenv("GEMINI_API_KEY", "")
        if not gemini_key:
            raise HTTPException(500, "Gemini API key not configured")

        genai.configure(api_key=gemini_key)
        model = genai.GenerativeModel("gemini-2.5-flash")

        prompt = f"""Look at this image and identify each distinct sellable item.
Return a JSON object with a list of bounding boxes for each item.
Coordinates are normalized 0-1000 in format [ymin, xmin, ymax, xmax].
For each item include a brief 1-3 word label.
Return at most {MULTI_ITEM_CAP} items, prioritizing the most clearly visible.

Format:
{{
  "items": [
    {{"box_2d": [100, 200, 400, 500], "label": "blue jeans"}},
    {{"box_2d": [50, 600, 350, 900], "label": "red shirt"}}
  ]
}}

Return ONLY the JSON object, no markdown, no other text."""

        pil_image = Image.open(_io.BytesIO(contents))
        if pil_image.mode != "RGB":
            pil_image = pil_image.convert("RGB")

        response = model.generate_content(
            [prompt, pil_image],
            generation_config={"max_output_tokens": 4000}
        )
        raw = (response.text or "").strip()
        # Strip code fences
        import re as _re
        raw = _re.sub(r"^```[a-zA-Z]*\n?", "", raw, flags=_re.IGNORECASE)
        raw = _re.sub(r"```$", "", raw).strip()
        import json
        data = json.loads(raw)
        items = (data.get("items") or [])[:MULTI_ITEM_CAP]

        # Get original image dimensions for scaling
        from PIL import Image
        import io as _io
        img = Image.open(_io.BytesIO(contents))
        w, h = img.size

        # No longer caching original server-side — frontend re-uploads it on /finalize
        import uuid
        upload_id = str(uuid.uuid4())

        # Convert normalized boxes to pixel coords
        out_items = []
        for i, it in enumerate(items):
            box = it.get("box_2d") or [0, 0, 0, 0]
            if len(box) != 4:
                continue
            ymin, xmin, ymax, xmax = box
            # Gemini returns 0-1000 normalized
            px = {
                "x": int(xmin * w / 1000),
                "y": int(ymin * h / 1000),
                "w": int((xmax - xmin) * w / 1000),
                "h": int((ymax - ymin) * h / 1000),
            }
            out_items.append({
                "id": i,
                "label": it.get("label", f"Item {i+1}"),
                "box": px,
                "selected": True
            })

        return {
            "ok": True,
            "upload_id": upload_id,
            "image_width": w,
            "image_height": h,
            "items": out_items
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))


@app.post("/api/multi-scan/finalize")
async def multi_scan_finalize(request: Request):
    """Crop the original photo into N item photos and create scan groups for each."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        # Frontend now sends multipart form: file + items + condition (JSON-encoded)
        form = await request.form()
        file = form.get("file")
        if not file:
            raise HTTPException(400, "No file")
        condition = form.get("condition", "used")
        import json as _json
        items_json = form.get("items", "[]")
        items = _json.loads(items_json) if isinstance(items_json, str) else (items_json or [])
        if not items:
            raise HTTPException(400, "items required")

        # Re-check tier
        biz = supabase.table("businesses").select("scan_limit,is_admin").eq("id", business_id).limit(1).execute()
        b = biz.data[0] if biz.data else {}
        if not b.get("is_admin") and (b.get("scan_limit") or 0) < MULTI_ITEM_MIN_TIER_LIMIT:
            raise HTTPException(402, "Growth plan or higher required")

        original = await file.read()

        from PIL import Image, ImageOps
        import io as _io
        from datetime import datetime
        img = Image.open(_io.BytesIO(original))
        img = ImageOps.exif_transpose(img)

        import uuid
        session_id = str(uuid.uuid4())
        created_groups = []

        # Crop each item, upload as a new photo, create a group
        for idx, item in enumerate(items[:MULTI_ITEM_CAP]):
            box = item.get("box") or {}
            x = max(0, int(box.get("x", 0)))
            y = max(0, int(box.get("y", 0)))
            w = int(box.get("w", 0))
            h = int(box.get("h", 0))
            if w <= 10 or h <= 10:
                continue
            # Add 5% padding around the crop for context
            pad_x = int(w * 0.05)
            pad_y = int(h * 0.05)
            crop_box = (
                max(0, x - pad_x),
                max(0, y - pad_y),
                min(img.width, x + w + pad_x),
                min(img.height, y + h + pad_y)
            )
            cropped = img.crop(crop_box)
            # Resize if too big
            if cropped.width > 1600 or cropped.height > 1600:
                cropped.thumbnail((1600, 1600))

            buf = _io.BytesIO()
            cropped.convert("RGB").save(buf, format="JPEG", quality=88)
            crop_bytes = buf.getvalue()

            # Upload crop to part-photos
            dt = datetime.now()
            fn = f"{dt.strftime('%d%m%y')}_{dt.strftime('%H%M%S')}_{idx}.jpg"
            try:
                supabase.storage.from_("part-photos").upload(
                    path=fn,
                    file=crop_bytes,
                    file_options={"content-type": "image/jpeg", "upsert": "true"}
                )
            except Exception as _e:
                print(f"crop upload warn for {idx}: {_e}")
                continue

            # Create listing_group
            try:
                grp = supabase.table("listing_groups").insert({
                    "session_id": session_id,
                    "business_id": business_id,
                    "condition": condition,
                    "quantity": 1,
                    "status": "pending"
                }).execute()
                gid = grp.data[0]["id"]
                supabase.table("group_photos").insert({"group_id": gid, "photo_id": fn}).execute()
                created_groups.append(gid)
            except Exception as _e:
                print(f"group create warn for {idx}: {_e}")
                continue

        return {"ok": True, "session_id": session_id, "groups_created": len(created_groups)}
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))


# ─── Demo Scanner (Landing Page) ──────────────────────────────

DEMO_RATE_LIMIT = 3  # scans per IP per 24h
_demo_counts = {}    # in-memory: {ip: [(timestamp, ...), ...]}


def _demo_get_ip(request: Request) -> str:
    fwd = request.headers.get("x-forwarded-for", "")
    if fwd:
        return fwd.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def _demo_check_quota(ip: str) -> tuple[bool, int]:
    """Return (allowed, remaining_after_use)."""
    from datetime import datetime, timezone, timedelta
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(hours=24)
    # Prune old entries
    if ip in _demo_counts:
        _demo_counts[ip] = [t for t in _demo_counts[ip] if t > cutoff]
    used = len(_demo_counts.get(ip, []))
    if used >= DEMO_RATE_LIMIT:
        return False, 0
    return True, DEMO_RATE_LIMIT - used - 1


@app.get("/api/demo-scan/quota")
async def demo_quota(request: Request):
    ip = _demo_get_ip(request)
    from datetime import datetime, timezone, timedelta
    cutoff = datetime.now(timezone.utc) - timedelta(hours=24)
    if ip in _demo_counts:
        _demo_counts[ip] = [t for t in _demo_counts[ip] if t > cutoff]
    used = len(_demo_counts.get(ip, []))
    return {"used": used, "limit": DEMO_RATE_LIMIT, "remaining": max(0, DEMO_RATE_LIMIT - used)}


@app.post("/api/demo-scan")
async def demo_scan(request: Request):
    """Lightweight one-shot Gemini scan for the landing page demo. No auth, no DB, no scanner queue."""
    ip = _demo_get_ip(request)
    allowed, _remaining = _demo_check_quota(ip)
    if not allowed:
        raise HTTPException(429, "Free demo limit reached. Sign up for unlimited scans.")
    try:
        form = await request.form()
        file = form.get("file")
        if not file:
            raise HTTPException(400, "No file uploaded")
        contents = await file.read()
        if len(contents) > 10 * 1024 * 1024:
            raise HTTPException(413, "Image too large (10MB max)")

        import google.generativeai as genai
        from PIL import Image, ImageOps
        import io as _io
        import json as _json

        gemini_key = os.getenv("GEMINI_API_KEY", "")
        if not gemini_key:
            raise HTTPException(500, "Service temporarily unavailable")

        genai.configure(api_key=gemini_key)
        model = genai.GenerativeModel("gemini-2.5-flash")

        try:
            img = Image.open(_io.BytesIO(contents))
            img = ImageOps.exif_transpose(img)
            if img.mode != "RGB":
                img = img.convert("RGB")
            # Resize for speed
            if img.width > 1600 or img.height > 1600:
                img.thumbnail((1600, 1600))
        except Exception as _img_err:
            raise HTTPException(400, f"Could not read image: {_img_err}")

        prompt = """You are a reseller's pricing assistant. Look at this photo and identify the most valuable item for resale.

Return ONLY a JSON object with this exact shape:
{
  "title": "Item title under 80 chars suitable for an eBay listing",
  "category": "Broad category like Electronics, Clothing, Tools, Home, Toys, Books, Auto, Sports, Collectibles, Health, or Other",
  "condition_guess": "new" or "used",
  "estimated_price_used": <integer dollar amount>,
  "estimated_price_new": <integer dollar amount>,
  "confidence": "high", "medium", or "low",
  "notes": "One short sentence with your pricing reasoning"
}

Pricing rules:
- Base prices on REAL eBay sold-listing values for similar items
- Both used and new prices required (estimate even if condition is unclear)
- Use 0 if you genuinely cannot price it
- Be realistic - undercut retail by 20-40% for used resale value

Return ONLY the JSON object, no markdown, no other text."""

        response = model.generate_content(
            [prompt, img],
            generation_config={
                "max_output_tokens": 600,
                "temperature": 0.0,
                "response_mime_type": "application/json"
            }
        )
        raw = (response.text or "").strip()
        import re as _re
        # Strip code fences
        raw = _re.sub(r"^```[a-zA-Z]*\n?", "", raw, flags=_re.IGNORECASE)
        raw = _re.sub(r"```$", "", raw).strip()
        # If there's extra text around the JSON, extract just the JSON object
        match = _re.search(r"\{[\s\S]*\}", raw)
        if match:
            raw = match.group(0)
        try:
            data = _json.loads(raw)
        except Exception as _parse_err:
            print(f"DEMO SCAN PARSE FAIL — raw response was:\n{raw[:500]}")
            raise HTTPException(500, "Could not parse AI response")

        # Record the use after success
        from datetime import datetime, timezone
        _demo_counts.setdefault(ip, []).append(datetime.now(timezone.utc))

        return {
            "ok": True,
            "title": data.get("title", "Unknown Item"),
            "category": data.get("category", "Other"),
            "condition_guess": data.get("condition_guess", "used"),
            "estimated_price_used": int(data.get("estimated_price_used", 0) or 0),
            "estimated_price_new": int(data.get("estimated_price_new", 0) or 0),
            "confidence": data.get("confidence", "medium"),
            "notes": data.get("notes", ""),
            "remaining": max(0, DEMO_RATE_LIMIT - len(_demo_counts.get(ip, [])))
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, f"Scan failed: {str(e)[:200]}")


# ─── Intake Inventory ──────────────────────────────────────────

@app.get("/api/intake")
async def list_intake(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        result = supabase.table("intake_purchases").select("*").eq("business_id", business_id).order("created_at", desc=True).execute()
        items = result.data or []
        for it in items:
            it["photo_url"] = f"{SUPABASE_URL}/storage/v1/object/public/part-photos/{it['photo_id']}" if it.get("photo_id") else ""
        return {"items": items}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/intake")
async def create_intake(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        row = {
            "business_id": business_id,
            "item_name": body.get("item_name") or "",
            "purchase_date": body.get("purchase_date") or None,
            "purchase_price": body.get("purchase_price") or None,
            "sku_code": body.get("sku_code") or "",
            "notes": body.get("notes") or "",
        }
        result = supabase.table("intake_purchases").insert(row).execute()
        return {"ok": True, "item": result.data[0] if result.data else None}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.patch("/api/intake/{intake_id}")
async def update_intake(intake_id: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        allowed = ("item_name", "purchase_date", "purchase_price", "sku_code", "notes", "photo_id")
        updates = {k: (v if v not in ("", None) else None) for k, v in body.items() if k in allowed}
        if not updates:
            raise HTTPException(400, "No valid fields")
        from datetime import datetime, timezone
        updates["updated_at"] = datetime.now(timezone.utc).isoformat()
        supabase.table("intake_purchases").update(updates).eq("id", intake_id).eq("business_id", business_id).execute()
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/api/intake/{intake_id}")
async def delete_intake(intake_id: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        supabase.table("intake_purchases").delete().eq("id", intake_id).eq("business_id", business_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/intake/upload-photo")
async def upload_intake_photo(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        form = await request.form()
        file = form.get("file")
        if not file:
            raise HTTPException(400, "No file")
        contents = await file.read()
        if len(contents) > 10 * 1024 * 1024:
            raise HTTPException(413, "Image too large (10MB max)")
        from datetime import datetime
        from PIL import Image, ImageOps
        import io as _io
        img = Image.open(_io.BytesIO(contents))
        img = ImageOps.exif_transpose(img)
        if img.mode != "RGB":
            img = img.convert("RGB")
        if img.width > 1600 or img.height > 1600:
            img.thumbnail((1600, 1600))
        buf = _io.BytesIO()
        img.save(buf, format="JPEG", quality=88)
        crop_bytes = buf.getvalue()
        dt = datetime.now()
        fn = f"intake_{dt.strftime('%d%m%y_%H%M%S')}_{business_id[:8]}.jpg"
        supabase.storage.from_("part-photos").upload(
            path=fn,
            file=crop_bytes,
            file_options={"content-type": "image/jpeg", "upsert": "true"}
        )
        return {"ok": True, "photo_id": fn, "photo_url": f"{SUPABASE_URL}/storage/v1/object/public/part-photos/{fn}"}
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))


# ─── Expenses ──────────────────────────────────────────────────

EXPENSE_CATEGORIES = [
    "Cost of Goods", "Shipping & Postage", "Supplies & Packaging",
    "Fees & Subscriptions", "Gas & Travel", "Food & Meals",
    "Software & Tools", "Other"
]

CATEGORY_COLORS = {
    "Cost of Goods":       "#ef4444",
    "Shipping & Postage":  "#f97316",
    "Supplies & Packaging":"#eab308",
    "Fees & Subscriptions":"#22c55e",
    "Gas & Travel":        "#3b82f6",
    "Food & Meals":        "#a855f7",
    "Software & Tools":    "#06b6d4",
    "Other":               "#6b7280",
}

@app.get("/api/expenses")
async def list_expenses(request: Request, period: str = "month", category: str = ""):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        from datetime import datetime, timezone, timedelta
        now = datetime.now(timezone.utc)
        date_filter = None
        if period == "day":
            date_filter = (now - timedelta(days=1)).date()
        elif period == "week":
            date_filter = (now - timedelta(weeks=1)).date()
        elif period == "month":
            date_filter = (now - timedelta(days=30)).date()
        elif period == "3month":
            date_filter = (now - timedelta(days=90)).date()
        elif period == "6month":
            date_filter = (now - timedelta(days=180)).date()
        elif period == "12month":
            date_filter = (now - timedelta(days=365)).date()

        q = supabase.table("expenses").select("*").eq("business_id", business_id)
        if date_filter:
            q = q.gte("expense_date", str(date_filter))
        if category:
            q = q.eq("category", category)
        result = q.order("expense_date", desc=True).execute()
        items = result.data or []

        # Summary by category
        summary = {}
        for cat in EXPENSE_CATEGORIES:
            summary[cat] = {"total": 0, "count": 0, "color": CATEGORY_COLORS.get(cat, "#6b7280")}
        for it in items:
            cat = it.get("category", "Other")
            if cat not in summary:
                summary[cat] = {"total": 0, "count": 0, "color": "#6b7280"}
            summary[cat]["total"] += float(it.get("amount") or 0)
            summary[cat]["count"] += 1

        total = sum(float(i.get("amount") or 0) for i in items)
        return {"items": items, "summary": summary, "total": total, "period": period}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/expenses")
async def create_expense(request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        row = {
            "business_id": business_id,
            "amount": float(body.get("amount") or 0),
            "category": body.get("category") or "Other",
            "merchant": body.get("merchant") or "",
            "notes": body.get("notes") or "",
            "expense_date": body.get("expense_date") or str(__import__("datetime").date.today()),
            "source": body.get("source") or "manual",
        }
        result = supabase.table("expenses").insert(row).execute()
        return {"ok": True, "item": result.data[0] if result.data else None}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.patch("/api/expenses/{expense_id}")
async def update_expense(expense_id: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        allowed = ("amount","category","merchant","notes","expense_date","receipt_photo_id")
        updates = {k: v for k, v in body.items() if k in allowed}
        if "amount" in updates:
            updates["amount"] = float(updates["amount"] or 0)
        from datetime import datetime, timezone
        updates["updated_at"] = datetime.now(timezone.utc).isoformat()
        supabase.table("expenses").update(updates).eq("id", expense_id).eq("business_id", business_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.delete("/api/expenses/{expense_id}")
async def delete_expense(expense_id: str, request: Request):
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        supabase.table("expenses").delete().eq("id", expense_id).eq("business_id", business_id).execute()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/expenses/scan-receipt")
async def scan_receipt(request: Request):
    """Gemini one-shot receipt scan — returns extracted data for user to confirm."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        form = await request.form()
        file = form.get("file")
        if not file:
            raise HTTPException(400, "No file")
        contents = await file.read()
        from PIL import Image, ImageOps
        import io as _io, json as _json, re as _re
        img = Image.open(_io.BytesIO(contents))
        img = ImageOps.exif_transpose(img)
        if img.mode != "RGB":
            img = img.convert("RGB")
        if img.width > 1600 or img.height > 1600:
            img.thumbnail((1600, 1600))
        buf = _io.BytesIO()
        img.save(buf, format="JPEG", quality=88)

        from google import genai as _genai
        import base64
        genai_client = _genai.Client(api_key=os.getenv("GEMINI_API_KEY",""))
        img_b64 = base64.b64encode(buf.getvalue()).decode()

        prompt = """You are a receipt scanning assistant. Extract the following from this receipt image and return ONLY raw JSON, no other text:
{
  "merchant": "store or vendor name",
  "amount": <total amount as number, no currency symbol>,
  "expense_date": "YYYY-MM-DD",
  "category": "one of: Cost of Goods, Shipping & Postage, Supplies & Packaging, Fees & Subscriptions, Gas & Travel, Food & Meals, Software & Tools, Other",
  "notes": "brief description of what was purchased"
}
If you cannot read a field clearly, use null. For category, make your best guess based on the merchant and items."""

        from google.genai import types as _gtypes
        response = genai_client.models.generate_content(
            model="models/gemini-2.5-flash",
            contents=[
                _gtypes.Content(parts=[
                    _gtypes.Part(inline_data=_gtypes.Blob(mime_type="image/jpeg", data=buf.getvalue())),
                    _gtypes.Part(text=prompt)
                ])
            ],
            config=_gtypes.GenerateContentConfig(
                max_output_tokens=300,
                temperature=0.0,
            )
        )
        raw = (response.text or "").strip()
        if not raw:
            # Try getting text from parts
            try:
                raw = response.candidates[0].content.parts[0].text.strip()
            except Exception:
                raw = ""
        raw = _re.sub(r"^```[a-zA-Z]*\n?", "", raw, flags=_re.IGNORECASE)
        raw = _re.sub(r"\n?```$", "", raw).strip()
        match = _re.search(r"\{[\s\S]*\}", raw)
        if match:
            raw = match.group(0)
        if not raw:
            return {"ok": True, "data": {"merchant": "", "amount": None, "expense_date": "", "category": "Other", "notes": "Could not read receipt automatically — please fill in manually"}}
        from json_repair import repair_json
        data = _json.loads(repair_json(raw))
        return {"ok": True, "data": data}
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(500, str(e))


@app.post("/api/expenses/toggle-cogs")
async def toggle_cogs_import(request: Request):
    """Toggle COGS import from scans. Records timestamp when enabled."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        body = await request.json()
        enabled = bool(body.get("enabled", False))
        from datetime import datetime, timezone
        updates = {"cogs_import_enabled": enabled}
        if enabled:
            updates["cogs_import_since"] = datetime.now(timezone.utc).isoformat()
        supabase.table("businesses").update(updates).eq("id", business_id).execute()
        return {"ok": True, "enabled": enabled}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/expenses/export")
async def export_expenses(request: Request, period: str = "month", fmt: str = "csv"):
    """Export expenses as CSV."""
    business_id = require_auth(request)
    if not business_id:
        raise HTTPException(401, "Not authenticated")
    try:
        from datetime import datetime, timezone, timedelta
        import csv, io as _io
        now = datetime.now(timezone.utc)
        date_filter = None
        if period != "all":
            days = {"day":1,"week":7,"month":30,"3month":90,"6month":180,"12month":365}.get(period,30)
            date_filter = (now - timedelta(days=days)).date()
        q = supabase.table("expenses").select("*").eq("business_id", business_id)
        if date_filter:
            q = q.gte("expense_date", str(date_filter))
        result = q.order("expense_date", desc=True).execute()
        items = result.data or []

        buf = _io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Date","Merchant","Category","Amount","Notes","Source"])
        for it in items:
            writer.writerow([
                it.get("expense_date",""),
                it.get("merchant",""),
                it.get("category",""),
                f"${float(it.get('amount') or 0):.2f}",
                it.get("notes",""),
                it.get("source","manual"),
            ])
        total = sum(float(i.get("amount") or 0) for i in items)
        writer.writerow([])
        writer.writerow(["TOTAL","","",f"${total:.2f}","",""])

        from fastapi.responses import Response
        return Response(
            content=buf.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=expenses_{period}.csv"}
        )
    except Exception as e:
        raise HTTPException(500, str(e))

