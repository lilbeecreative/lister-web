content = open('main.py').read()

# 1. Add /auction route
old_route = '@app.get("/v2", response_class=HTMLResponse)'
new_route = '''@app.get("/auction", response_class=HTMLResponse)
async def auction_page(request: Request):
    import os
    with open(os.path.join(os.path.dirname(__file__), "templates", "auction.html")) as f:
        html = f.read()
    return HTMLResponse(content=html, headers={
        "Content-Security-Policy": "default-src * blob: data:; script-src * blob: data: 'unsafe-inline' 'unsafe-eval'; style-src * 'unsafe-inline'; img-src * blob: data:;"
    })

@app.get("/v2", response_class=HTMLResponse)'''

if old_route in content:
    content = content.replace(old_route, new_route)
    print('✅ /auction route added')
else:
    print('❌ route anchor not found')

# 2. Add deep research endpoint and excel export before the closing section
deep_endpoint = '''
# ── API: AUCTION DEEP RESEARCH ───────────────────────────────── #

class DeepResearch(BaseModel):
    title: str
    current_value: float = 0

@app.post("/api/auction/deep-research")
async def deep_research(body: DeepResearch):
    import os, asyncio, json
    from concurrent.futures import ThreadPoolExecutor
    import google.generativeai as genai
    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(400, "GEMINI_API_KEY not set")
    genai.configure(api_key=gemini_key)
    model = genai.GenerativeModel("gemini-2.5-flash")
    prompt = f"""You are an expert industrial equipment appraiser.
Do a thorough market research on this auction item: "{body.title}"
Current estimate: ${body.current_value}

Research:
1. Search eBay sold listings for this exact item or very similar models
2. Check industrial equipment dealers
3. Consider age, condition (assume working used), brand reputation

Return ONLY a JSON object:
{{"your_value": 5000, "notes": "Sold for $4,500-$6,000 on eBay in 2024, industrial dealers list at $8,000+"}}

your_value must be an integer. Return ONLY the JSON, no markdown."""

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)
    try:
        response = await loop.run_in_executor(executor, lambda: model.generate_content(prompt, generation_config={"max_output_tokens": 500}))
        raw = response.text.strip().replace('```json','').replace('```','').strip()
        data = json.loads(raw)
        return data
    except Exception as e:
        raise HTTPException(500, str(e))


# ── API: EXCEL EXPORT ─────────────────────────────────────────── #

@app.post("/api/auction/export-excel")
async def export_excel(request: Request):
    import io, json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    body = await request.json()
    items = body.get('items', [])
    name = body.get('name', 'Auction Scan')

    wb = Workbook()
    ws = wb.active
    ws.title = name[:31]

    # Header style
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1A1F2E")
    hv_fill = PatternFill("solid", fgColor="412402")
    hv_font = Font(color="FAC775", bold=True)
    normal_fill = PatternFill("solid", fgColor="1E2535")

    headers = ["Lot", "Title", "Description", "Est. Value", "Notes", "Deep Scan", "Watchlisted"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for item in items:
        val = item.get('your_value', 0) or 0
        row = [
            str(item.get('lot', '')),
            str(item.get('title', '')),
            str(item.get('description', '')),
            f"${val:,}",
            str(item.get('notes', '')),
            'Yes' if item.get('_deep') else '',
            'Yes' if item.get('_watch') else '',
        ]
        ws.append(row)
        r = ws.max_row
        if val >= 500:
            for col in range(1, 8):
                ws.cell(row=r, column=col).fill = hv_fill
                ws.cell(row=r, column=col).font = hv_font
        else:
            for col in range(1, 8):
                ws.cell(row=r, column=col).fill = normal_fill

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import datetime
    fn = f"auction_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"}
    )

'''

anchor = '\n# ── API: AUCTION ──'
if anchor in content:
    content = content.replace(anchor, deep_endpoint + anchor)
    print('✅ Deep research + Excel endpoints added')
else:
    print('❌ Auction anchor not found')

open('main.py', 'w').write(content)
