content = open('main.py').read()

endpoint = '''
@app.post("/api/auction/research-export")
async def research_export(request: Request):
    import io, json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    form = await request.form()
    items = json.loads(form.get("items", "[]"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Deep Research"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E2535")
    amber_fill = PatternFill("solid", fgColor="412402")
    green_fill = PatternFill("solid", fgColor="052E16")
    alt_fill = PatternFill("solid", fgColor="161B28")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="center")
    thin = Border(
        bottom=Side(style="thin", color="2D3348"),
        right=Side(style="thin", color="2D3348")
    )

    headers = ["Lot", "Title", "Original Value", "Revised Value", "Confidence", "Recommendation", "Notes", "eBay Search"]
    col_widths = [8, 45, 15, 15, 12, 16, 40, 20]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 22

    for ri, item in enumerate(items, 2):
        row_data = [
            item.get("lot", ""),
            item.get("title", ""),
            item.get("original_value", 0),
            item.get("revised_value", 0),
            item.get("confidence", "").capitalize(),
            item.get("recommendation", "").capitalize(),
            item.get("rec_reason") or item.get("image_notes", ""),
            "View eBay Sold"
        ]
        rec = item.get("recommendation", "").lower()
        fill = green_fill if rec == "buy" else amber_fill if rec == "watch" else (alt_fill if ri % 2 == 0 else None)

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill
            cell.border = thin
            cell.alignment = wrap if ci in (2, 7) else center
            if ci in (3, 4) and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0'
            if ci == 8 and item.get("ebay_search"):
                ws.cell(row=ri, column=ci).hyperlink = item["ebay_search"]
                ws.cell(row=ri, column=ci).font = Font(color="4A9EFF", underline="single")
        ws.row_dimensions[ri].height = 20

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=deep_research.xlsx"}
    )

'''

anchor = '# ── API: AUCTION DEEP RESEARCH ───────────────────────────────── #'
if anchor in content:
    content = content.replace(anchor, endpoint + anchor)
    open('main.py', 'w').write(content)
    print('done')
else:
    print('not found')
