content = open('main.py').read()

new_endpoints = '''
# ── API: AUCTION DEEP RESEARCH ───────────────────────────────── #

class DeepResearch(BaseModel):
    title: str
    current_value: float = 0

@app.post("/api/auction/deep-research")
async def deep_research(body: DeepResearch):
    import os, asyncio, json
    from concurrent.futures import ThreadPoolExecutor
    import google.generativeai as genai
    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        raise HTTPException(400, "GEMINI_API_KEY not set")
    genai.configure(api_key=gemini_key)
    model = genai.GenerativeModel("gemini-2.5-flash")
    prompt = f"""You are an expert industrial equipment appraiser.
Research this auction item thoroughly: "{body.title}"
Current estimate: ${body.current_value}

Check eBay sold listings, industrial dealers, and recent auction results.
Assume working used condition.

Return ONLY a JSON object (no markdown):
{{"your_value": 5000, "notes": "Sold $4,500-$6,000 on eBay 2024"}}

your_value must be an integer."""

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)
    try:
        response = await loop.run_in_executor(executor, lambda: model.generate_content(prompt, generation_config={"max_output_tokens": 300}))
        raw = response.text.strip().replace("```json","").replace("```","").strip()
        data = json.loads(raw)
        return data
    except Exception as e:
        raise HTTPException(500, str(e))


# ── API: EXCEL EXPORT ─────────────────────────────────────────── #

@app.post("/api/auction/export-excel")
async def export_excel(request: Request):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    body = await request.json()
    items = body.get("items", [])
    name = body.get("name", "Auction Scan")

    wb = Workbook()
    ws = wb.active
    ws.title = name[:31]

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1A1F2E")
    hv_fill = PatternFill("solid", fgColor="5C3A00")
    hv_font = Font(color="FAC775", bold=True)

    headers = ["Lot", "Title", "Est. Value", "Notes", "Deep Scan", "Watchlisted"]
    ws.append(headers)
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        val = int(item.get("your_value", 0) or 0)
        row = [
            str(item.get("lot", "")),
            str(item.get("title", "")),
            f"${val:,}",
            str(item.get("notes", "")),
            "Yes" if item.get("_deep") else "",
            "Yes" if item.get("_watch") else "",
        ]
        ws.append(row)
        if val >= 500:
            r = ws.max_row
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = hv_fill
                ws.cell(row=r, column=col).font = hv_font

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import datetime
    fn = f"auction_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}"}
    )

'''

anchor = '# ── API: PDF AUCTION SCAN ─────────────────────────────────────── #'
if anchor in content:
    content = content.replace(anchor, new_endpoints + anchor)
    open('main.py', 'w').write(content)
    print('done')
else:
    print('not found')
